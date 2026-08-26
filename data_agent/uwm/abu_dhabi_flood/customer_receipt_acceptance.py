"""Fail-closed acceptance of the private Abu Dhabi SWMM customer receipt.

The workbook is a customer communication artifact, not an authority by
itself.  This module verifies that each claimed delivery has a resolvable
file, matching SHA-256, and the minimum provenance metadata before marking an
issue as accepted.  It never opens any model admission gate.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA = "gwm.abu_dhabi_flood.customer_swmm_engineering_receipt_acceptance.v1"
REGISTER_SCHEMA = "gwm.abu_dhabi_flood.customer_swmm_engineering_issue_register.v1"
SHEET_NAME = "客户回执"
HEADER_ROW = 4
DATA_START_ROW = 5

REQUIRED_HEADERS = {
    "问题编号",
    "客户交付状态",
    "客户文件名/私有路径",
    "数据来源单位/责任方",
    "版本或快照编号",
    "有效时间/事件范围",
    "水平坐标系（CRS）",
    "垂直基准",
    "单位",
    "质量标识/缺测说明",
    "许可或交付授权",
    "SHA-256",
    "客户备注",
}

_SPATIAL_TERMS = (
    "高程",
    "地形",
    "管网",
    "管径",
    "流向",
    "端点",
    "出水口",
    "汇水区",
    "进水口",
    "泵站",
    "道路",
    "建筑",
)
_VERTICAL_TERMS = ("高程", "地形", "管底", "潮位", "水位", "垂直")
_PUBLIC_PROXY_TERMS = (
    "open-meteo",
    "open meteo",
    "公开代理",
    "public proxy",
    "synthetic",
    "合成测试",
    "测试替代",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_paths(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[;；\n]+", value) if part.strip()]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _resolve_path(raw: str, data_roots: list[Path]) -> tuple[Path | None, str | None]:
    candidate = Path(raw).expanduser()
    if candidate.is_absolute():
        return (candidate.resolve(), None) if candidate.is_file() else (None, "file_not_found")
    for root in data_roots:
        resolved = (root / candidate).resolve()
        if resolved.is_file():
            return resolved, None
    return None, "file_not_found"


def _requires_crs(issue: dict[str, Any]) -> bool:
    title = _text(issue.get("title"))
    return any(term in title for term in _SPATIAL_TERMS)


def _requires_vertical_datum(issue: dict[str, Any]) -> bool:
    title = _text(issue.get("title"))
    return any(term in title for term in _VERTICAL_TERMS)


def _required_metadata(issue: dict[str, Any]) -> list[str]:
    required = [
        "数据来源单位/责任方",
        "版本或快照编号",
        "有效时间/事件范围",
        "单位",
        "质量标识/缺测说明",
        "许可或交付授权",
    ]
    if _requires_crs(issue):
        required.append("水平坐标系（CRS）")
    if _requires_vertical_datum(issue):
        required.append("垂直基准")
    return required


def _row_map(sheet: Any) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        header = _text(sheet.cell(HEADER_ROW, column).value)
        if header:
            mapping[header] = column
    missing = sorted(REQUIRED_HEADERS.difference(mapping))
    if missing:
        raise ValueError("customer_receipt_headers_missing:" + ",".join(missing))
    return mapping


def _read_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError("customer_receipt_sheet_missing")
    sheet = workbook[SHEET_NAME]
    columns = _row_map(sheet)
    rows: list[dict[str, Any]] = []
    for row_number in range(DATA_START_ROW, sheet.max_row + 1):
        values = {
            header: sheet.cell(row_number, column).value for header, column in columns.items()
        }
        if not any(_text(value) for value in values.values()):
            continue
        values["_row_number"] = row_number
        rows.append(values)
    if not rows:
        raise ValueError("customer_receipt_rows_missing")
    return rows


def _validate_one(
    issue: dict[str, Any],
    row: dict[str, Any],
    data_roots: list[Path],
) -> dict[str, Any]:
    issue_id = _text(issue.get("issue_id"))
    delivery_status = _text(row.get("客户交付状态"))
    paths = _split_paths(_text(row.get("客户文件名/私有路径")))
    expected_hashes = [part.lower() for part in _split_paths(_text(row.get("SHA-256")))]
    missing_metadata = [field for field in _required_metadata(issue) if not _text(row.get(field))]
    reasons: list[str] = []
    files: list[dict[str, Any]] = []
    if delivery_status in ("", "待客户提供"):
        reasons.append("customer_delivery_not_received")
    elif delivery_status == "部分提供":
        reasons.append("customer_delivery_partial")
    elif delivery_status == "不适用":
        reasons.append("not_applicable_requires_project_approval")
    elif delivery_status != "已提供待验收":
        reasons.append("unknown_customer_delivery_status")
    if not paths:
        reasons.append("customer_file_path_missing")
    if not expected_hashes:
        reasons.append("sha256_missing")
    if paths and expected_hashes and len(paths) != len(expected_hashes):
        reasons.append("file_and_sha256_count_mismatch")
    for index, raw_path in enumerate(paths):
        resolved, error = _resolve_path(raw_path, data_roots)
        item: dict[str, Any] = {"submitted_path": raw_path, "exists": resolved is not None}
        if error:
            item["error"] = error
            reasons.append(f"file_{index + 1}_not_found")
        else:
            actual_hash = _sha256(resolved)
            item["resolved_path"] = str(resolved)
            item["size_bytes"] = resolved.stat().st_size
            item["sha256"] = actual_hash
            if index >= len(expected_hashes):
                reasons.append(f"file_{index + 1}_sha256_missing")
            elif actual_hash != expected_hashes[index]:
                item["expected_sha256"] = expected_hashes[index]
                item["hash_match"] = False
                reasons.append(f"file_{index + 1}_sha256_mismatch")
            else:
                item["hash_match"] = True
        files.append(item)
    if missing_metadata:
        reasons.append("metadata_missing")
    customer_note = _text(row.get("客户备注"))
    provenance_text = " ".join(
        _text(row.get(field)) for field in ("数据来源单位/责任方", "客户备注", "许可或交付授权")
    ).casefold()
    if any(term in provenance_text for term in _PUBLIC_PROXY_TERMS):
        reasons.append("public_proxy_cannot_close_customer_issue")
    if delivery_status == "不适用" and not customer_note:
        reasons.append("not_applicable_reason_missing")
    accepted = not reasons
    if accepted:
        review_status = "验收通过"
        disposition = "accepted_customer_authoritative_delivery"
    elif delivery_status in ("已提供待验收", "部分提供", "不适用"):
        review_status = "退回补充"
        disposition = "return_for_customer_completion"
    else:
        review_status = "待审核"
        disposition = "waiting_for_customer_delivery"
    return {
        "issue_id": issue_id,
        "priority": _text(issue.get("priority")),
        "title": _text(issue.get("title")),
        "workbook_row": row.get("_row_number"),
        "customer_delivery_status": delivery_status or "未填写",
        "project_review_status": review_status,
        "disposition": disposition,
        "accepted": accepted,
        "required_metadata": _required_metadata(issue),
        "missing_metadata": missing_metadata,
        "files": files,
        "reasons": reasons,
        "customer_notes_present": bool(customer_note),
    }


def accept_customer_receipt(
    workbook_path: Path,
    register_path: Path,
    data_roots: list[Path] | None = None,
) -> dict[str, Any]:
    """Return a deterministic, fail-closed acceptance receipt."""

    workbook_path = workbook_path.expanduser().resolve()
    register_path = register_path.expanduser().resolve()
    register = json.loads(register_path.read_text(encoding="utf-8"))
    if register.get("schema") != REGISTER_SCHEMA:
        raise ValueError("unexpected_engineering_issue_register_schema")
    issues = register.get("issues", [])
    if not issues:
        raise ValueError("engineering_issue_register_empty")
    rows = _read_rows(workbook_path)
    by_id: dict[str, dict[str, Any]] = {}
    for row in rows:
        issue_id = _text(row.get("问题编号"))
        if not issue_id:
            raise ValueError(f"customer_receipt_issue_id_missing:row_{row['_row_number']}")
        if issue_id in by_id:
            raise ValueError(f"customer_receipt_duplicate_issue_id:{issue_id}")
        by_id[issue_id] = row
    expected_ids = {str(issue["issue_id"]) for issue in issues}
    actual_ids = set(by_id)
    missing_ids = sorted(expected_ids - actual_ids)
    unexpected_ids = sorted(actual_ids - expected_ids)
    if missing_ids or unexpected_ids:
        details = []
        if missing_ids:
            details.append("missing=" + ",".join(missing_ids))
        if unexpected_ids:
            details.append("unexpected=" + ",".join(unexpected_ids))
        raise ValueError("customer_receipt_issue_ids_invalid:" + ";".join(details))
    roots = [root.expanduser().resolve() for root in (data_roots or [workbook_path.parent])]
    results = [_validate_one(issue, by_id[str(issue["issue_id"])], roots) for issue in issues]
    accepted_count = sum(item["accepted"] for item in results)
    review_counts: dict[str, int] = {}
    for item in results:
        review_counts[item["project_review_status"]] = (
            review_counts.get(item["project_review_status"], 0) + 1
        )
    p0_results = [item for item in results if item["priority"] == "P0"]
    p0_accepted = sum(item["accepted"] for item in p0_results)
    return {
        "schema": SCHEMA,
        "version": "2026-08-23",
        "status": "customer_receipt_accepted"
        if accepted_count == len(results)
        else "customer_receipt_requires_action",
        "source": {
            "workbook": str(workbook_path),
            "register": str(register_path),
            "data_roots": [str(root) for root in roots],
            "customer_rows_copied_to_public_repository": False,
        },
        "summary": {
            "issue_count": len(results),
            "accepted_issue_count": accepted_count,
            "blocked_issue_count": len(results) - accepted_count,
            "p0_issue_count": len(p0_results),
            "p0_accepted_issue_count": p0_accepted,
            "review_status_counts": review_counts,
        },
        "admission": {
            "customer_authoritative_delivery_complete": accepted_count == len(results),
            "engineering_calibration_admitted": False,
            "traditional_model_admitted": False,
            "gwm_training_admitted": False,
            "hybrid_planner_admitted": False,
            "city_scale_prediction_claim_allowed": False,
            "admission_note": (
                "Receipt acceptance only verifies delivery evidence; independent engineering "
                "calibration and event validation remain required."
            ),
        },
        "issues": results,
    }


def render_acceptance_markdown(payload: dict[str, Any]) -> str:
    summary = payload["summary"]
    lines = [
        "# 阿布扎比城市暴雨内涝世界模型",
        "## SWMM工程数据客户回执自动验收结果",
        "",
        (
            f"状态：**{payload['status']}**。共 {summary['issue_count']} 项，"
            f"验收通过 {summary['accepted_issue_count']} 项，仍需处理 "
            f"{summary['blocked_issue_count']} 项。"
        ),
        "",
        "自动验收仅核对客户填写内容、文件存在性、SHA-256和最低元数据，不等同于工程校准或模型准入。工程校准、传统模型正式运行、GWM训练、混合规划器和城市级预测声明仍保持关闭。",
        "",
        "| 问题编号 | 优先级 | 项目审核状态 | 结论 | 原因 |",
        "|---|---|---|---|---|",
    ]
    for item in payload["issues"]:
        reasons = "、".join(item["reasons"]) if item["reasons"] else "无"
        verdict = "通过" if item["accepted"] else "未通过"
        lines.append(
            f"| {item['issue_id']} | {item['priority']} | "
            f"{item['project_review_status']} | {verdict} | {reasons} |"
        )
    lines.extend(
        [
            "",
            "## 准入状态",
            "",
            "```json",
            json.dumps(payload["admission"], ensure_ascii=False, indent=2),
            "```",
            "",
        ]
    )
    return "\n".join(lines)
