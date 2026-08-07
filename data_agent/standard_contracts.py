"""Readers for auditable Ningxia data-model baseline catalogs.

The two Ningxia workbooks are the supplied inventory and field baseline. They
are sufficient to identify datasets and propose field-level bindings before a
real FileGDB arrives. Runtime ingestion still verifies physical field types,
lengths, CRS, geometry, values and quality for each dataset. A separately
reviewed ``ea_standard`` catalog remains supported for explicit administrative
publication workflows, but it is not a prerequisite for starting the agent.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

# Older bundles use the *_candidate labels. They remain readable as runtime
# baselines; the per-dataset schema and quality gates make the real decision.
RUNTIME_BASELINE_AUTHORITIES = frozenset(
    {
        "nx_workbook_baseline",
        "nx_project_baseline",
        "ea_standard",
        "ea_analysis_candidate",
        "standard_candidate",
        "screenshot_candidate",
    }
)

# These are exact business-name mappings backed by the compiled natural-resource
# standard, not fuzzy guesses. Broad items such as "社会经济数据" intentionally
# remain unresolved until their concrete table or file schema is known.
INVENTORY_STANDARD_CONTRACT_HINTS = {
    "行政区划界线": ("XZQJX",),
    "应急避难场所": ("YJBNA",),
}

STANDARD_GEOMETRY_TYPES = {
    "XZQJX": "LineString",
    "YJBNA": "Polygon",
}

STANDARD_PRIMARY_KEYS = {
    "XZQJX": "BSM",
    "YJBNA": "FEATID",
}


def _now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_runtime_baseline_authority(value: Any) -> bool:
    return _text(value) in RUNTIME_BASELINE_AUTHORITIES


def normalize_identifier(value: Any) -> str:
    """Normalize field/layer names only for comparison, never for storage."""
    return re.sub(r"[\s_\-（）()：:./\\]+", "", _text(value)).lower()


def _header_row(rows: list[tuple[Any, ...]], required: set[str]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        columns = {_text(value): position for position, value in enumerate(row) if _text(value)}
        if required.issubset(columns):
            return index, columns
    raise ValueError(f"workbook header not found: {sorted(required)}")


def _workbook_rows(path: Path, sheet: str) -> list[tuple[Any, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"sheet not found: {sheet}")
        return [tuple(row) for row in workbook[sheet].iter_rows(values_only=True)]
    finally:
        workbook.close()


def _load_workbook_sheet_names(path: str | Path) -> list[str]:
    """Return sheet names without treating a workbook as a data contract."""
    from openpyxl import load_workbook

    workbook = load_workbook(Path(path).expanduser().resolve(), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_shp_contract_catalog(path: str | Path) -> dict[str, Any]:
    """Parse the SHP workbook into contracts keyed by the layer code."""
    source = Path(path).expanduser().resolve()
    summary_rows = _workbook_rows(source, "数据表汇总")
    detail_rows = _workbook_rows(source, "字段明细")
    summary_at, summary_columns = _header_row(
        summary_rows, {"图层/数据表代码", "中文名称", "几何类型", "完整性/核验说明"}
    )
    detail_at, detail_columns = _header_row(
        detail_rows, {"图层/数据表代码", "字段序号", "字段名称", "字段类别"}
    )
    contracts: dict[str, dict[str, Any]] = {}
    for row in summary_rows[summary_at + 1 :]:
        code = _text(row[summary_columns["图层/数据表代码"]])
        if not code:
            continue
        note = _text(row[summary_columns["完整性/核验说明"]])
        contracts[code] = {
            "code": code,
            "name": _text(row[summary_columns["中文名称"]]),
            "geometry_type": _text(row[summary_columns["几何类型"]]),
            "field_count": int(row[summary_columns["可辨字段数"]] or 0)
            if "可辨字段数" in summary_columns
            else None,
            "source_photos": _text(row[summary_columns["来源照片"]]).split("、")
            if "来源照片" in summary_columns
            else [],
            "name_source": _text(row[summary_columns["名称来源"]])
            if "名称来源" in summary_columns
            else "",
            "completeness_note": note,
            "fields": [],
            "authority": "nx_workbook_baseline",
            "publication_gate": "per_dataset_runtime_validation",
            "requires_source_schema_verification": True,
        }
    for row in detail_rows[detail_at + 1 :]:
        code = _text(row[detail_columns["图层/数据表代码"]])
        field_name = _text(row[detail_columns["字段名称"]])
        if not code or not field_name or code not in contracts:
            continue
        contracts[code]["fields"].append(
            {
                "ordinal": int(row[detail_columns["字段序号"]] or 0),
                "name": field_name,
                "category": _text(row[detail_columns["字段类别"]]),
                "source_photo": _text(row[detail_columns["来源照片"]])
                if "来源照片" in detail_columns
                else "",
                "source": "SHP workbook screenshot",
            }
        )
    for contract in contracts.values():
        contract["fields"].sort(key=lambda item: item["ordinal"])
        contract["candidate_fields"] = [item["name"] for item in contract["fields"]]
        contract["required_fields"] = [
            item["name"]
            for item in contract["fields"]
            if item.get("category") != "系统/几何字段"
        ] or list(contract["candidate_fields"])
        contract["recommended_fields"] = [
            item["name"]
            for item in contract["fields"]
            if item.get("category") == "系统/几何字段"
        ]
        contract["field_categories"] = {
            item["name"]: item["category"] for item in contract["fields"]
        }
    return {
        "schema_version": "gda.standard-contract-catalog.v1",
        "generated_at": _now(),
        "source_workbooks": [source.name],
        "authority": "nx_workbook_baseline",
        "runtime_baseline_ready": True,
        "publication_gate": "per_dataset_runtime_validation",
        "contracts": contracts,
    }


def load_contract_catalog(path: str | Path) -> dict[str, Any]:
    """Load either an Excel discovery workbook or a generated JSON catalog."""
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() in {".xlsx", ".xlsm", ".xltx"}:
        return load_shp_contract_catalog(source)
    payload = json.loads(source.read_text(encoding="utf-8"))
    if "contracts" in payload:
        return payload
    # Legacy aliases JSON remains supported as a field-only catalog.
    aliases_payload = payload.get("field_aliases") or {
        key: value for key, value in payload.items() if isinstance(value, dict)
    }
    contracts = {
        code: {
            "code": code,
            "name": code,
            "geometry_type": "",
            "fields": [
                {
                    "ordinal": index,
                    "name": name,
                    "aliases": list(values),
                    "category": "",
                }
                for index, (name, values) in enumerate(fields.items(), 1)
            ],
            "candidate_fields": list(fields.keys()),
            "authority": "nx_workbook_baseline",
            "publication_gate": "per_dataset_runtime_validation",
            "requires_source_schema_verification": True,
        }
        for code, fields in aliases_payload.items()
    }
    return {"schema_version": "gda.standard-contract-catalog.v1", "contracts": contracts, **payload}


def aliases_from_catalog(catalog: dict[str, Any]) -> dict[str, dict[str, set[str]]]:
    """Build matching aliases without changing the workbook's displayed names."""
    result: dict[str, dict[str, set[str]]] = {}
    for code, contract in (catalog.get("contracts") or {}).items():
        result[code] = {}
        for field in contract.get("fields") or []:
            name = _text(field.get("code")) or _text(field.get("name"))
            display_name = _text(field.get("name"))
            if name:
                result[code][name] = {
                    name.lower(),
                    normalize_identifier(name),
                    display_name.lower() if display_name else "",
                    normalize_identifier(display_name) if display_name else "",
                    *(str(value).lower() for value in field.get("aliases", [])),
                } - {""}
    return result


def load_inventory_summary(path: str | Path) -> dict[str, Any]:
    """Keep the first workbook's data-list evidence alongside contracts."""
    source = Path(path).expanduser().resolve()
    rows = _workbook_rows(source, "客户提供数据清单梳理")
    header_at, columns = _header_row(rows, {"数据名称", "数据大类", "数据小类"})
    items = []
    for row in rows[header_at + 1 :]:
        name = _text(row[columns["数据名称"]])
        if not name:
            continue
        items.append(
            {
                "category": _text(row[columns["数据大类"]]),
                "subcategory": _text(row[columns["数据小类"]]),
                "name": name,
                "use_case": _text(row[columns["城市体检应用场景数据作用"]])
                if "城市体检应用场景数据作用" in columns
                else "",
                "format": _text(row[columns["数据格式"]]) if "数据格式" in columns else "",
                "note": _text(row[columns["说明"]]) if "说明" in columns else "",
                "source": "customer data inventory workbook",
            }
        )
    return {"source_workbook": source.name, "item_count": len(items), "items": items}


def _inventory_field_category(name: str) -> str:
    lowered = _text(name).lower()
    if lowered in {
        "objectid",
        "shape",
        "shape_length",
        "shape_area",
        "shape_length_new",
        "shape_area_new",
    }:
        return "系统/几何字段"
    if any(token in lowered for token in ("代码", "编码", "标识", "编号", "id")):
        return "标识编码字段"
    if any(token in lowered for token in ("时间", "日期", "年份", "年")):
        return "时态字段"
    if any(token in lowered for token in ("数据源", "涉密", "来源")):
        return "来源与安全字段"
    return "业务属性字段"


def _inventory_contract(
    code: str, name: str, sheet: str, fields: list[str], geometry_type: str = ""
) -> dict[str, Any]:
    unique_fields = list(dict.fromkeys(_text(value) for value in fields if _text(value)))
    field_rows = [
        {
            "ordinal": index,
            "name": field,
            "category": _inventory_field_category(field),
            "source_sheet": sheet,
            "source": "inventory workbook field sheet",
        }
        for index, field in enumerate(unique_fields, 1)
    ]
    candidate_fields = [item["name"] for item in field_rows]
    required_fields = [
        item["name"] for item in field_rows if item["category"] != "系统/几何字段"
    ] or candidate_fields
    return {
        "code": code,
        "name": name,
        "geometry_type": geometry_type,
        "field_count": len(unique_fields),
        "source_sheets": [sheet],
        "name_source": "inventory workbook field sheet",
        "completeness_note": "字段行来自宁夏数据清单工作簿；真实文件到达时复核物理 schema。",
        "fields": field_rows,
        "candidate_fields": candidate_fields,
        "required_fields": required_fields,
        "recommended_fields": [
            item["name"] for item in field_rows if item["category"] == "系统/几何字段"
        ],
        "field_categories": {item["name"]: item["category"] for item in field_rows},
        "authority": "nx_workbook_baseline",
        "publication_gate": "per_dataset_runtime_validation",
        "requires_source_schema_verification": True,
    }


def load_inventory_field_baseline(path: str | Path) -> dict[str, Any]:
    """Extract the field rows embedded in the inventory workbook."""
    source = Path(path).expanduser().resolve()
    contracts: dict[str, dict[str, Any]] = {}

    def first_header(sheet: str, required: set[str]) -> list[str] | None:
        rows = _workbook_rows(source, sheet)
        for row in rows:
            values = [_text(value) for value in row if _text(value)]
            if required.issubset(values):
                return values
        return None

    known_sheets = set(_load_workbook_sheet_names(source))
    if "2024年土地利用现状调查数据" in known_sheets:
        fields = first_header("2024年土地利用现状调查数据", {"标识码", "要素代码", "图斑编号"})
        if fields:
            contracts["DLTB"] = _inventory_contract(
                "DLTB", "地类图斑", "2024年土地利用现状调查数据", fields, "Polygon"
            )
    if "2025年不动产-ZRZ" in known_sheets:
        fields = first_header("2025年不动产-ZRZ", {"实体标识码", "标识码"})
        if fields:
            contracts["ZRZ"] = _inventory_contract(
                "ZRZ", "自然幢", "2025年不动产-ZRZ", fields, "Polygon"
            )
    planning_sheet = "自治区级国土空间总体规划数据（3条）"
    if planning_sheet in known_sheets:
        rows = _workbook_rows(source, planning_sheet)
        planning_codes = {
            "CZJSSYXPJJG": "城镇建设适宜性评价结果",
            "NYSCSYXPJJG": "农业生产适宜性评价结果",
            "STBHZYXPJJG": "生态保护重要性评价结果",
        }
        for index, row in enumerate(rows[:-1]):
            values = [_text(value) for value in row if _text(value)]
            code = next(
                (
                    item
                    for item in planning_codes
                    if any(item in value for value in values)
                ),
                None,
            )
            if not code:
                continue
            fields = [_text(value) for value in rows[index + 1] if _text(value)]
            if len(fields) >= 2:
                contracts[code] = _inventory_contract(
                    code, planning_codes[code], planning_sheet, fields, "Polygon"
                )
    county_sheet = "县级国土空间规划（仅治理历史文化保护线 中心城区黄绿蓝紫线规）"
    if county_sheet in known_sheets:
        fields = first_header(county_sheet, {"标识码", "要素代码", "行政区代码"})
        if fields:
            contracts["ZXCQ"] = _inventory_contract(
                "ZXCQ", "中心城区规划分区", county_sheet, fields, "Polygon"
            )
    building_sheet = "银川市城市存量成果FWJZ2024"
    if building_sheet in known_sheets:
        fields = first_header(building_sheet, {"建筑编码", "基底面积", "建筑高度"})
        if fields:
            contracts["FWJZ"] = _inventory_contract(
                "FWJZ", "房屋建筑数据", building_sheet, fields, "Polygon"
            )
    metric_sheet = "银川市城市存量成果SQCPG2025"
    if metric_sheet in known_sheets:
        fields = first_header(metric_sheet, {"空间单元代码", "空间单元名称", "建筑密度"})
        if fields:
            contracts["SQCPG"] = _inventory_contract(
                "SQCPG", "城市空间品质指标", metric_sheet, fields, ""
            )
    return {
        "source_workbook": source.name,
        "contracts": contracts,
        "dataset_count": len(contracts),
        "field_count": sum(len(item.get("fields") or []) for item in contracts.values()),
    }


def _standard_data_type(value: Any) -> str | None:
    normalized = _text(value).lower()
    if normalized in {"char", "varchar", "text", "string"}:
        return "string"
    if normalized in {"int", "integer", "long", "short"}:
        return "integer"
    if normalized in {"float", "double", "number", "numeric", "decimal"}:
        return "number"
    if normalized in {"date", "datetime", "timestamp"}:
        return normalized
    if normalized in {"bool", "boolean"}:
        return "boolean"
    return normalized or None


def load_standard_document_contracts(
    path: str | Path, codes: set[str]
) -> dict[str, dict[str, Any]]:
    """Load selected complete field tables from the compiled standard document."""
    import yaml

    source = Path(path).expanduser().resolve()
    payload = yaml.safe_load(source.read_text(encoding="utf-8")) or {}
    selected: dict[str, dict[str, Any]] = {}
    evidence: dict[str, list[dict[str, Any]]] = {}
    for table in payload.get("field_tables") or []:
        code = _text(table.get("table_code")).upper()
        source_fields = [
            field
            for field in table.get("fields") or []
            if _text(field.get("code"))
        ]
        if code not in codes or not source_fields:
            continue
        evidence.setdefault(code, []).append(
            {
                "module": table.get("module"),
                "table_name_cn": table.get("table_name_cn"),
                "caption_raw": table.get("caption_raw"),
                "section_path": table.get("section_path") or [],
                "field_count": len(source_fields),
            }
        )
        current = selected.get(code)
        if current and len(current.get("fields") or []) >= len(source_fields):
            continue
        fields = []
        for ordinal, field in enumerate(source_fields, 1):
            field_code = _text(field.get("code"))
            field_name = _text(field.get("name_cn"))
            constraint = _text(field.get("constraint")).upper()
            length_text = _text(field.get("length"))
            precision_text = _text(field.get("decimal"))
            fields.append(
                {
                    "ordinal": int(field.get("seq") or ordinal),
                    "code": field_code,
                    "name": field_name or field_code,
                    "aliases": list(dict.fromkeys([field_code, field_name]))
                    if field_name
                    else [field_code],
                    "data_type": _standard_data_type(field.get("dtype")),
                    "type_source": "natural_resource_standard_document",
                    "length": int(length_text) if length_text.isdigit() else None,
                    "precision": int(precision_text) if precision_text.isdigit() else None,
                    "required": constraint == "M",
                    "nullable": constraint != "M",
                    "primary_key": field_code == STANDARD_PRIMARY_KEYS.get(code),
                    "unique": field_code == STANDARD_PRIMARY_KEYS.get(code),
                    "unit": None,
                    "value_domain": {
                        "name": _text(field.get("domain")),
                        "values": [],
                    }
                    if _text(field.get("domain"))
                    else None,
                    "quality_rules": {
                        "constraint": constraint or None,
                        "note": _text(field.get("note")) or None,
                    },
                    "authority": "nx_project_baseline",
                }
            )
        required = [field["code"] for field in fields if field["required"]]
        recommended = [field["code"] for field in fields if not field["required"]]
        selected[code] = {
            "code": code,
            "name": _text(table.get("table_name_cn")).replace("属性结构描述表", ""),
            "module": _text(table.get("module")),
            "geometry_type": STANDARD_GEOMETRY_TYPES.get(code),
            "primary_key": STANDARD_PRIMARY_KEYS.get(code),
            "required_fields": required,
            "recommended_fields": recommended,
            "fields": fields,
            "standard_document_evidence": evidence[code],
            "sources": {
                "standard_documents": [payload.get("source_file") or source.name],
                "compiled_standard_catalog": source.name,
            },
            "authority": "nx_project_baseline",
            "review_status": "runtime_validation",
            "publication_gate": "per_dataset_runtime_validation",
            "requires_source_schema_verification": True,
            "runtime_baseline_ready": True,
            "auto_publish": False,
            "field_completeness": {
                "field_count": len(fields),
                "required_count": len(required),
                "recommended_count": len(recommended),
                "missing_type_count": sum(not field["data_type"] for field in fields),
                "missing_length_count": sum(field["length"] is None for field in fields),
                "missing_precision_count": sum(
                    field["precision"] is None for field in fields
                ),
                "missing_domain_count": sum(
                    field["value_domain"] is None for field in fields
                ),
            },
        }
    return selected


def build_catalog(
    shp_workbook: str | Path,
    inventory_workbook: str | Path | None = None,
) -> dict[str, Any]:
    catalog = load_shp_contract_catalog(shp_workbook)
    if inventory_workbook:
        catalog["data_inventory"] = load_inventory_summary(inventory_workbook)
        inventory_baseline = load_inventory_field_baseline(inventory_workbook)
        for code, source_contract in inventory_baseline["contracts"].items():
            if code not in catalog["contracts"]:
                catalog["contracts"][code] = _workbook_contract_as_runtime_baseline(
                    source_contract
                )
            else:
                catalog["contracts"][code]["inventory_field_baseline"] = source_contract[
                    "fields"
                ]
                catalog["contracts"][code]["inventory_source_sheets"] = source_contract.get(
                    "source_sheets", []
                )
        catalog["inventory_field_baseline"] = inventory_baseline
    return catalog


def _workbook_contract_as_runtime_baseline(contract: dict[str, Any]) -> dict[str, Any]:
    """Convert one Excel field inventory row into a runtime contract object."""
    source_fields = contract.get("fields") or []
    required = list(contract.get("required_fields") or contract.get("candidate_fields") or [])
    required_set = set(required)
    fields = [
        {
            "ordinal": item.get("ordinal"),
            "code": item.get("name"),
            "name": item.get("name"),
            "aliases": [item.get("name")],
            "category": item.get("category"),
            "data_type": None,
            "type_source": "runtime_source_schema",
            "length": None,
            "precision": None,
            "required": item.get("name") in required_set,
            "nullable": item.get("name") not in required_set,
            "primary_key": False,
            "unique": False,
            "unit": None,
            "value_domain": None,
            "quality_rules": {},
            "source_photo": item.get("source_photo"),
            "source_sheet": item.get("source_sheet"),
            "source": item.get("source"),
            "authority": "nx_workbook_baseline",
        }
        for item in source_fields
    ]
    geometry_map = {"面": "Polygon", "点": "Point", "线": "LineString"}
    return {
        **contract,
        "geometry_type": geometry_map.get(
            contract.get("geometry_type"), contract.get("geometry_type")
        ),
        "required_fields": required,
        "recommended_fields": list(contract.get("recommended_fields") or []),
        "fields": fields,
        "authority": "nx_workbook_baseline",
        "review_status": "runtime_validation",
        "publication_gate": "per_dataset_runtime_validation",
        "requires_source_schema_verification": True,
        "runtime_baseline_ready": True,
        "auto_publish": False,
        "field_completeness": {
            "field_count": len(fields),
            "required_count": sum(item.get("required") is True for item in fields),
            "recommended_count": len(contract.get("recommended_fields") or []),
            "missing_type_count": len(fields),
            "missing_length_count": len(fields),
            "missing_precision_count": len(fields),
            "missing_domain_count": len(fields),
        },
    }


def merge_workbook_baseline(
    contracts: dict[str, dict[str, Any]],
    shp_workbook: str | Path | None,
    inventory_workbook: str | Path | None = None,
) -> dict[str, Any]:
    """Merge both Ningxia field workbooks without overwriting source evidence."""
    coverage = {
        "dataset_count": 0,
        "field_count": 0,
        "overlap_count": 0,
        "inventory_field_dataset_count": 0,
        "inventory_field_count": 0,
        "inventory_overlap_count": 0,
    }
    if shp_workbook:
        workbook_catalog = load_shp_contract_catalog(shp_workbook)
        coverage["dataset_count"] = len(workbook_catalog.get("contracts") or {})
        for code, source_contract in (workbook_catalog.get("contracts") or {}).items():
            coverage["field_count"] += len(source_contract.get("fields") or [])
            if code not in contracts:
                contracts[code] = _workbook_contract_as_runtime_baseline(source_contract)
                continue
            coverage["overlap_count"] += 1
            target = contracts[code]
            target["workbook_field_baseline"] = source_contract.get("fields") or []
            target["workbook_field_categories"] = source_contract.get("field_categories") or {}
            target["workbook_completeness_note"] = source_contract.get("completeness_note")
            target["workbook_source_photos"] = source_contract.get("source_photos") or []
            target["runtime_baseline_ready"] = True
    if inventory_workbook:
        inventory_catalog = load_inventory_field_baseline(inventory_workbook)
        coverage["inventory_field_dataset_count"] = inventory_catalog["dataset_count"]
        coverage["inventory_field_count"] = inventory_catalog["field_count"]
        for code, source_contract in (inventory_catalog.get("contracts") or {}).items():
            if code not in contracts:
                contracts[code] = _workbook_contract_as_runtime_baseline(source_contract)
                continue
            coverage["inventory_overlap_count"] += 1
            target = contracts[code]
            target["inventory_field_baseline"] = source_contract.get("fields") or []
            target["inventory_field_categories"] = source_contract.get("field_categories") or {}
            target["inventory_source_sheets"] = source_contract.get("source_sheets") or []
            target["inventory_completeness_note"] = source_contract.get("completeness_note")
            target["runtime_baseline_ready"] = True
    return coverage


def _sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).expanduser().resolve().open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_json(path: str | Path) -> dict[str, Any]:
    return json.loads(Path(path).expanduser().resolve().read_text(encoding="utf-8"))


def _load_csv_rows(path: str | Path | None) -> list[dict[str, str]]:
    if not path:
        return []
    with Path(path).expanduser().resolve().open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _infer_field_type(field: str, rule: dict[str, Any]) -> tuple[str | None, str]:
    """Infer only a machine-useful candidate type; never call it authoritative."""
    explicit = _text(rule.get("type")).lower()
    if explicit:
        return explicit, "rule_candidate"
    pattern = _text(rule.get("pattern"))
    if pattern == "^[0-9]{4}$":
        return "string", "pattern_candidate"
    if pattern == "^[0-9]{8}$":
        return "string", "pattern_candidate"
    return None, "missing_authoritative_definition"


def compile_standard_contract_catalog(
    *,
    role_contracts_path: str | Path,
    field_aliases_path: str | Path | None = None,
    value_domains_path: str | Path | None = None,
    field_catalog_path: str | Path | None = None,
    ea_table_comparison_path: str | Path | None = None,
    ea_logical_comparison_path: str | Path | None = None,
    standard_docx_catalog_path: str | Path | None = None,
    shp_workbook: str | Path | None = None,
    inventory_workbook: str | Path | None = None,
    standard_version: str | None = None,
) -> dict[str, Any]:
    """Compile the Ningxia runtime baseline used by the ingestion control plane.

    The workbooks and EA comparison exports establish dataset and field
    coverage. Missing physical details are checked against each real source
    file at ingest time; they no longer prevent the whole Windows installation
    from starting. Explicit administrative publication can still call
    :func:`validate_contract_catalog` with ``require_authoritative=True``.
    """
    role_payload = _load_json(role_contracts_path)
    aliases_payload = _load_json(field_aliases_path) if field_aliases_path else {}
    domain_payload = _load_json(value_domains_path) if value_domains_path else {}
    field_rows = {
        _text(row.get("field_name")): row
        for row in _load_csv_rows(field_catalog_path)
        if _text(row.get("field_name"))
    }
    ea_rows = _load_csv_rows(ea_table_comparison_path)
    logical_rows = _load_csv_rows(ea_logical_comparison_path)
    aliases = aliases_payload.get("field_aliases") or {}
    domains = domain_payload.get("domains") or {}
    contracts: dict[str, dict[str, Any]] = {}
    source_paths = [
        role_contracts_path,
        field_aliases_path,
        value_domains_path,
        field_catalog_path,
        ea_table_comparison_path,
        ea_logical_comparison_path,
        standard_docx_catalog_path,
        shp_workbook,
        inventory_workbook,
    ]
    source_artifacts = [
        {
            "name": Path(path).name,
            "sha256": _sha256_file(path),
            "portable_locator": "deployment_config_source",
        }
        for path in source_paths
        if path and Path(path).expanduser().exists()
    ]
    for role_id, role in (role_payload.get("roles") or {}).items():
        required = list(dict.fromkeys(role.get("required_fields") or []))
        recommended = [
            field
            for field in dict.fromkeys(role.get("recommended_fields") or [])
            if field not in required
        ]
        rules = role.get("field_rules") or {}
        ordered_fields = (
            required
            + recommended
            + [field for field in rules if field not in required + recommended]
        )
        table_specs = role.get("standard_tables") or []
        for table_spec in table_specs:
            code = _text(table_spec.get("table_code"))
            if not code:
                continue
            fields = []
            missing_type = 0
            for ordinal, field in enumerate(ordered_fields, 1):
                rule = rules.get(field) or {}
                data_type, type_source = _infer_field_type(field, rule)
                registry = field_rows.get(field) or {}
                if not data_type:
                    missing_type += 1
                alias = _text(aliases.get(field)) or _text(registry.get("field_alias_zh"))
                domain_name = _text(rule.get("domain"))
                fields.append(
                    {
                        "ordinal": ordinal,
                        "code": field,
                        "name": alias,
                        "aliases": [value for value in [field, alias] if value],
                        "data_type": data_type,
                        "type_source": type_source,
                        "length": rule.get("length"),
                        "precision": rule.get("precision"),
                        "required": field in required,
                        "nullable": field not in required,
                        "primary_key": field == (role.get("twm_binding") or {}).get("object_id"),
                        "unique": field == (role.get("twm_binding") or {}).get("object_id"),
                        "unit": rule.get("unit"),
                        "value_domain": {
                            "name": domain_name,
                            "values": domains.get(domain_name, []),
                        }
                        if domain_name
                        else None,
                        "quality_rules": rule,
                        "source_registry": registry or None,
                        "authority": "nx_project_baseline",
                    }
                )
            ea_evidence = [
                {
                    key: row.get(key, "")
                    for key in (
                        "match_mode",
                        "ea_object_id",
                        "ea_name",
                        "ea_alias",
                        "ea_path",
                        "ea_fields",
                        "matched_fields",
                        "field_coverage_pct",
                        "missing_fields",
                        "type_mismatches",
                        "length_mismatches",
                    )
                }
                for row in ea_rows + logical_rows
                if _text(row.get("standard_table")).lower() == code.lower()
            ]
            contracts[code] = {
                "code": code,
                "name": _text(table_spec.get("table_alias_zh")),
                "role_id": role_id,
                "module": _text(table_spec.get("module")),
                "geometry_type": role.get("geometry_type"),
                "spatial_reference": {
                    "required": "CGCS2000",
                    "srid": None,
                    "status": "pending_authoritative_confirmation",
                },
                "time_model": {
                    "status": "candidate",
                    "keys": [
                        key for key in ((role.get("twm_binding") or {}).get("temporal_key"),) if key
                    ],
                },
                "primary_key": (role.get("twm_binding") or {}).get("object_id"),
                "required_fields": required,
                "recommended_fields": recommended,
                "fields": fields,
                "quality_rules": {"role_field_rules": rules},
                "ea_evidence": ea_evidence,
                "sources": {
                    "standard_documents": role_payload.get("source_documents") or [],
                    "standard_package": (
                        Path(_text(role_payload.get("source_package"))).name
                        if _text(role_payload.get("source_package"))
                        else None
                    ),
                    "ea_comparison_is_summary_only": True,
                },
                "authority": "nx_project_baseline",
                "review_status": "runtime_validation",
                "publication_gate": "per_dataset_runtime_validation",
                "auto_publish": False,
                "runtime_baseline_ready": True,
                "field_completeness": {
                    "field_count": len(fields),
                    "required_count": len(required),
                    "recommended_count": len(recommended),
                    "missing_type_count": missing_type,
                    "missing_length_count": sum(item["length"] is None for item in fields),
                    "missing_precision_count": sum(item["precision"] is None for item in fields),
                    "missing_domain_count": sum(item["value_domain"] is None for item in fields),
                },
            }
    standard_document_contracts = (
        load_standard_document_contracts(
            standard_docx_catalog_path,
            {
                code
                for codes in INVENTORY_STANDARD_CONTRACT_HINTS.values()
                for code in codes
            },
        )
        if standard_docx_catalog_path
        else {}
    )
    standard_document_overlap_count = 0
    for code, source_contract in standard_document_contracts.items():
        if code not in contracts:
            contracts[code] = source_contract
            continue
        standard_document_overlap_count += 1
        contracts[code]["standard_document_field_baseline"] = source_contract["fields"]
        contracts[code]["standard_document_evidence"] = source_contract[
            "standard_document_evidence"
        ]
        contracts[code]["runtime_baseline_ready"] = True
    workbook_coverage = merge_workbook_baseline(
        contracts, shp_workbook, inventory_workbook
    )
    inventory = load_inventory_summary(inventory_workbook) if inventory_workbook else None
    inventory_resolution = []
    contract_code_lookup = {code.upper(): code for code in contracts}
    for item in (inventory or {}).get("items", []):
        item_name = _text(item.get("name"))
        tokens = {token.upper() for token in re.findall(r"[A-Za-z][A-Za-z0-9_]{1,}", item_name)}
        direct_contract_codes = list(
            dict.fromkeys(
                contract_code_lookup[token]
                for token in tokens
                if token in contract_code_lookup
            )
        )
        hinted_contract_codes = [
            code
            for marker, hinted_codes in INVENTORY_STANDARD_CONTRACT_HINTS.items()
            if marker in item_name
            for code in hinted_codes
            if code in contracts
        ]
        item_core = normalize_identifier(item_name)
        for generic in ("数据", "全区", "银川市", "自治区", "记录"):
            item_core = item_core.replace(generic, "")
        name_contract_codes = []
        if len(item_core) >= 2:
            for contract_code, contract in contracts.items():
                contract_core = normalize_identifier(contract.get("name"))
                for generic in ("数据", "全区", "银川市", "自治区", "面", "点", "线"):
                    contract_core = contract_core.replace(generic, "")
                if len(contract_core) >= 2 and (
                    contract_core in item_core or item_core in contract_core
                ):
                    name_contract_codes.append(contract_code)
        direct_contract_codes = list(
            dict.fromkeys(
                direct_contract_codes + hinted_contract_codes + name_contract_codes
            )
        )
        evidence = [
            row
            for row in ea_rows + logical_rows
            if _text(row.get("standard_table")).upper() in tokens
            or (
                len(_text(row.get("standard_title"))) >= 4
                and normalize_identifier(row.get("standard_title"))
                in normalize_identifier(item_name)
            )
        ]
        has_exact_evidence = bool(evidence)
        prefix_evidence = [
            row
            for row in ea_rows + logical_rows
            if any(
                len(token) >= 3
                and _text(row.get("standard_table"))
                and (
                    _text(row.get("standard_table")).upper().startswith(token)
                    or token.startswith(_text(row.get("standard_table")).upper())
                )
                for token in tokens
            )
        ]
        if not evidence and prefix_evidence:
            evidence = prefix_evidence
        evidence_codes = list(
            dict.fromkeys(
                _text(row.get("standard_table"))
                for row in evidence
                if _text(row.get("standard_table"))
            )
        )
        codes = list(dict.fromkeys(direct_contract_codes + evidence_codes))
        contract_codes = (
            direct_contract_codes
            if direct_contract_codes
            else list(dict.fromkeys(code for code in evidence_codes if code in contracts))
        )
        if contract_codes:
            status = "baseline_contract"
            action = "接入真实文件时核验字段类型、长度、坐标系和质量"
        elif prefix_evidence and not has_exact_evidence:
            status = "ambiguous_candidate"
            action = "确认代码前缀对应的唯一标准对象后编译合同"
        elif evidence:
            status = "ea_aligned_baseline"
            action = "接入真实文件时核验字段类型、长度、坐标系和质量"
        else:
            status = "unresolved"
            action = "根据真实文件画像和标准正文人工确认标准对象"
        inventory_resolution.append(
            {
                **item,
                "status": status,
                "candidate_codes": codes,
                "contract_codes": contract_codes,
                "ea_evidence": [
                    {
                        key: row.get(key, "")
                        for key in (
                            "standard_table",
                            "standard_title",
                            "match_mode",
                            "ea_object_id",
                            "ea_name",
                            "ea_path",
                            "ea_fields",
                            "field_coverage_pct",
                        )
                    }
                    for row in evidence[:20]
                ],
                "required_action": action,
            }
        )
    inventory_coverage = {
        "items": inventory_resolution,
        "counts": {
            status: sum(item["status"] == status for item in inventory_resolution)
            for status in (
                "baseline_contract",
                "ea_aligned_baseline",
                "ambiguous_candidate",
                "unresolved",
            )
        },
    }
    catalog = {
        "schema_version": "gda.standard-contract-catalog.v2",
        "contract_id": "nx-natural-resource-standard-baseline",
        "generated_at": _now(),
        "standard_version": standard_version
        or role_payload.get("version")
        or "pending-confirmation",
        "authority": "nx_workbook_baseline",
        "review_status": "runtime_validation",
        "publication_gate": "per_dataset_runtime_validation",
        "production_ready": False,
        "runtime_baseline_ready": True,
        "provenance": {
            "source_artifacts": source_artifacts,
            "ea_comparison_warning": (
                "EA comparison CSVs are aggregate evidence, not attribute exports."
            ),
            "synthetic_field_catalog_warning": (
                "standard_field_catalog marks fields as synthetic/not_for_production."
            ),
        },
        "contracts": contracts,
        "field_registry": sorted(field_rows.values(), key=lambda row: row.get("field_name", "")),
        "value_domains": domains,
        "data_inventory": inventory,
        "inventory_resolution": inventory_coverage,
        "coverage": {
            "inventory_items": (inventory or {}).get("item_count", 0),
            "contract_dataset_count": len(contracts),
            "workbook_contract_dataset_count": workbook_coverage["dataset_count"],
            "workbook_field_count": workbook_coverage["field_count"],
            "workbook_overlap_count": workbook_coverage["overlap_count"],
            "inventory_field_contract_dataset_count": workbook_coverage[
                "inventory_field_dataset_count"
            ],
            "inventory_field_count": workbook_coverage["inventory_field_count"],
            "inventory_field_overlap_count": workbook_coverage[
                "inventory_overlap_count"
            ],
            "standard_document_contract_dataset_count": len(
                standard_document_contracts
            ),
            "standard_document_field_count": sum(
                len(contract.get("fields") or [])
                for contract in standard_document_contracts.values()
            ),
            "standard_document_overlap_count": standard_document_overlap_count,
            "unmapped_inventory": [
                item for item in inventory_resolution if item["status"] == "unresolved"
            ],
        },
    }
    catalog["coverage"]["unmapped_inventory_count"] = len(catalog["coverage"]["unmapped_inventory"])
    return catalog


def validate_contract_catalog(
    catalog: dict[str, Any], *, require_authoritative: bool = True
) -> list[str]:
    """Return deterministic blockers suitable for preflight and CI."""
    blockers: list[str] = []
    if catalog.get("schema_version") not in {
        "gda.standard-contract-catalog.v1",
        "gda.standard-contract-catalog.v2",
    }:
        blockers.append("contract schema version is not v1/v2")
    if require_authoritative and catalog.get("authority") != "ea_standard":
        blockers.append(
            f"catalog authority is {catalog.get('authority')!r}, expected 'ea_standard'"
        )
    for code, contract in (catalog.get("contracts") or {}).items():
        if require_authoritative and contract.get("authority") != "ea_standard":
            blockers.append(f"{code}: contract authority is not ea_standard")
        if not contract.get("fields"):
            blockers.append(f"{code}: no fields")
        for field in contract.get("fields") or []:
            if require_authoritative and not field.get("data_type"):
                blockers.append(f"{code}.{field.get('code')}: data_type missing")
    return blockers
