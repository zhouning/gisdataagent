from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from data_agent.offline_ingest import OfflineIngestStore
from data_agent.standard_contracts import load_shp_contract_catalog


async def _stream(data: bytes, block_size: int = 7):
    for offset in range(0, len(data), block_size):
        yield data[offset : offset + block_size]


@pytest.mark.asyncio
async def test_resumable_upload_commits_immutable_raw_asset(tmp_path):
    payload = b"FileGDB bundle placeholder\n" * 150000
    store = OfflineIngestStore(tmp_path / "lake")
    session = store.create_session(
        "宁夏_DLTB.gdb.zip",
        len(payload),
        chunk_size=1024 * 1024,
        expected_sha256=hashlib.sha256(payload).hexdigest(),
        asset_kind="filegdb_bundle",
        source_system="nx-survey",
    )
    session_id = session["session_id"]
    # Upload out of order, then verify the status is resumable.
    for index in (2, 0, 1):
        start = index * (1024 * 1024)
        await store.write_chunk(session_id, index, _stream(payload[start : start + (1024 * 1024)]))
    status = store.session_status(session_id)
    assert status["total_chunks"] == 4
    assert set(status["chunks"]) == {"0", "1", "2"}

    # Complete the remaining chunks after a simulated process interruption.
    for index in range(3, status["total_chunks"]):
        start = index * (1024 * 1024)
        await store.write_chunk(session_id, index, _stream(payload[start : start + (1024 * 1024)]))
    result = store.finalize_session(session_id, actor="test")
    assert result["status"] == "committed"
    committed = result["asset"]["path"]
    assert open(committed, "rb").read() == payload
    assert store.get_run(result["run_id"])["status"] == "succeeded"
    assert (tmp_path / "lake" / "runs" / result["run_id"] / "events.jsonl").exists()


@pytest.mark.asyncio
async def test_chunk_checksum_and_size_are_enforced(tmp_path):
    store = OfflineIngestStore(tmp_path / "lake")
    session = store.create_session("asset.tif", 4, chunk_size=1024 * 1024)
    with pytest.raises(ValueError, match="sha256 mismatch"):
        await store.write_chunk(
            session["session_id"], 0, _stream(b"data"), supplied_sha256="0" * 64
        )
    with pytest.raises(ValueError, match="chunk exceeds expected size"):
        await store.write_chunk(session["session_id"], 0, _stream(b"too-long"))


def test_local_scan_creates_quality_and_lineage_artifacts(tmp_path, monkeypatch):
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    (inbox / "elevation.tif").write_bytes(b"not-a-real-tiff")
    gdb = inbox / "DLTB.gdb"
    gdb.mkdir()
    (gdb / "a00000001.gdbtable").write_bytes(b"fixture")
    monkeypatch.setenv("GDA_LOCAL_INGEST_DIRS", str(inbox))
    store = OfflineIngestStore(tmp_path / "lake")
    result = store.scan_local_path(inbox, actor="tester")
    assert result["status"] == "blocked"
    assert result["asset_count"] == 2
    assert result["lineage"]
    for asset in result["assets"]:
        assert asset["raw_status"] == "committed"
        assert Path(asset["raw_path"]).exists()
    quality = json.loads(
        (tmp_path / "lake" / "runs" / result["run_id"] / "quality_report.json").read_text()
    )
    assert len(quality["items"]) == 2
    archive = store.export_diagnostics(result["run_id"])
    assert archive.exists()


def test_deep_quality_is_used_by_standardization_gate(tmp_path, monkeypatch):
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    (inbox / "broken.tif").write_bytes(b"not-a-real-tiff")
    monkeypatch.setenv("GDA_LOCAL_INGEST_DIRS", str(inbox))
    store = OfflineIngestStore(tmp_path / "lake")
    scan = store.scan_local_path(inbox, actor="tester")
    deep = store.run_deep_quality(scan["run_id"], actor="tester")
    assert deep["status"] == "blocked"
    parent = store.get_run(scan["run_id"])
    assert parent["deep_quality_run_id"] == deep["run_id"]
    with pytest.raises(ValueError, match="quality gate blocked"):
        store.create_standardization_plan(scan["run_id"], actor="tester", allow_review=True)


def test_standardization_plan_requires_quality_and_records_targets(tmp_path, monkeypatch):
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    archive_path = inbox / "DLTB.zip"
    with zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("DLTB.gdb/a.gdbtable", "fixture")
    monkeypatch.setenv("GDA_LOCAL_INGEST_DIRS", str(inbox))
    store = OfflineIngestStore(tmp_path / "lake")
    scan = store.scan_local_path(inbox, actor="tester")
    plan_run = store.create_standardization_plan(scan["run_id"], actor="tester")
    assert plan_run["status"] == "planned"
    plan = plan_run["standardization_plan"]
    assert plan["outputs"][0]["target_kind"] == "catalog_reference"
    assert (tmp_path / "lake" / "standardized" / scan["run_id"]).exists() is False
    assert (
        tmp_path / "lake" / "standardized" / plan_run["run_id"] / "standardization_plan.json"
    ).exists()
    execution = store.execute_standardization_plan(plan_run["run_id"], actor="tester")
    assert execution["status"] == "succeeded"
    assert (
        tmp_path / "lake" / "materialized" / plan_run["run_id"] / "materialization.json"
    ).exists()
    with pytest.raises(ValueError, match="accepted dataset baseline and quality evidence required"):
        store.create_ontology_binding(plan_run["run_id"], actor="tester")


def test_ontology_binding_accepts_quality_gated_nx_baseline(tmp_path):
    store = OfflineIngestStore(tmp_path / "lake")
    plan_id = "a" * 32
    plan_root = store.root / "standardized" / plan_id
    materialized_root = store.root / "materialized" / plan_id
    plan_root.mkdir(parents=True)
    materialized_root.mkdir(parents=True)
    (plan_root / "standardization_plan.json").write_text(
        json.dumps({"plan_id": plan_id, "status": "planned"}), encoding="utf-8"
    )
    (materialized_root / "materialization.json").write_text(
        json.dumps(
            {
                "status": "succeeded",
                "outputs": [
                    {
                        "target_id": f"standardized:{plan_id}:DLTB",
                        "target_kind": "postgis_or_geoparquet",
                        "target_path": "D:/GDA_DATA/governed/DLTB.parquet",
                        "target_sha256": "1" * 64,
                        "source_asset_id": "asset-dltb",
                        "canonical_dataset": "DLTB",
                        "execution_status": "succeeded",
                        "mapping": {
                            "status": "accepted",
                            "contract_authority": "nx_workbook_baseline",
                        },
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    result = store.create_ontology_binding(plan_id, actor="tester")

    assert result["status"] == "succeeded"
    binding = result["ontology_binding"]
    assert binding["production_eligible"] is True
    assert binding["bindings"][0]["mapping_authority"] == "nx_workbook_baseline"
    assert binding["bindings"][0]["target_sha256"] == "1" * 64


def test_rehearsal_raster_binding_does_not_treat_every_raster_as_orthophoto(tmp_path):
    store = OfflineIngestStore(tmp_path / "lake")
    plan_id = "b" * 32
    plan_root = store.root / "standardized" / plan_id
    materialized_root = store.root / "materialized" / plan_id
    plan_root.mkdir(parents=True)
    materialized_root.mkdir(parents=True)
    (plan_root / "standardization_plan.json").write_text(
        json.dumps({"plan_id": plan_id, "status": "planned"}), encoding="utf-8"
    )
    outputs = []
    for name in ("CLCD_2020.tif", "city_DOM.tif", "region_DEM.tif", "unknown.tif"):
        outputs.append(
            {
                "target_id": f"derived:{plan_id}:{name}",
                "target_kind": "cog_stac",
                "target_name": name,
                "target_path": f"D:/GDA_DATA/governed/{name}",
                "target_sha256": "2" * 64,
                "source_asset_id": f"asset:{name}",
                "execution_status": "succeeded",
                "mapping": {},
            }
        )
    (materialized_root / "materialization.json").write_text(
        json.dumps({"status": "succeeded", "outputs": outputs}), encoding="utf-8"
    )

    result = store.create_ontology_binding(
        plan_id, actor="tester", binding_mode="rehearsal"
    )

    binding = result["ontology_binding"]
    canonical = {item["canonical_dataset"] for item in binding["bindings"]}
    assert canonical == {"CLCD", "SZZSYX", "SZGCMX"}
    assert binding["skipped"] == [
        {"target_id": f"derived:{plan_id}:unknown.tif", "reason": "no_ontology_schema_candidate"}
    ]


def test_same_named_assets_cannot_overwrite_each_other_in_raw_zone(tmp_path, monkeypatch):
    inbox = tmp_path / "inbox"
    for folder, payload in (("a", b"first"), ("b", b"second")):
        target = inbox / folder
        target.mkdir(parents=True)
        (target / "同名影像.tif").write_bytes(payload)
    monkeypatch.setenv("GDA_LOCAL_INGEST_DIRS", str(inbox))
    result = OfflineIngestStore(tmp_path / "lake").scan_local_path(inbox)
    raw_paths = {item["raw_path"] for item in result["assets"]}
    assert len(raw_paths) == 2
    assert {Path(path).read_bytes() for path in raw_paths} == {b"first", b"second"}


def test_shapefile_sidecars_are_committed_as_one_bundle(tmp_path, monkeypatch):
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    shp = inbox / "roads.shp"
    shp.write_bytes(b"shape")
    shp.with_suffix(".dbf").write_bytes(b"attributes")
    shp.with_suffix(".prj").write_bytes(b"crs")
    monkeypatch.setenv("GDA_LOCAL_INGEST_DIRS", str(inbox))
    result = OfflineIngestStore(tmp_path / "lake").scan_local_path(inbox)
    asset = result["assets"][0]
    raw = Path(asset["raw_path"])
    assert sorted(path.suffix for path in raw.parent.glob(f"{raw.stem}.*")) == [
        ".dbf",
        ".prj",
        ".shp",
    ]
    assert asset["size"] == len(b"shape") + len(b"attributes") + len(b"crs")


def test_workbook_contract_is_runtime_baseline_and_exposes_field_categories(tmp_path, monkeypatch):
    workbook_path = tmp_path / "shp-contract.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "数据表汇总"
    summary.append(["说明"])
    summary.append(
        [
            "序号",
            "图层/数据表代码",
            "中文名称",
            "几何类型",
            "可辨字段数",
            "来源照片",
            "名称来源",
            "完整性/核验说明",
        ]
    )
    summary.append([1, "XXA", "学校（面）", "面", 3, "照片01", "截图显示", "截图字段列表完整可见"])
    details = workbook.create_sheet("字段明细")
    details.append(["说明"])
    details.append(
        [
            "总序号",
            "图层/数据表代码",
            "中文名称",
            "几何类型",
            "字段序号",
            "字段名称",
            "字段类别",
            "来源照片",
        ]
    )
    details.append([1, "XXA", "学校（面）", "面", 1, "OBJECTID", "系统/几何字段", "照片01"])
    details.append([2, "XXA", "学校（面）", "面", 2, "学校名称", "业务属性字段", "照片01"])
    details.append([3, "XXA", "学校（面）", "面", 3, "产生时间", "时态字段", "照片01"])
    workbook.create_sheet("分表字段清单")
    workbook.create_sheet("识别说明")
    workbook.save(workbook_path)

    catalog = load_shp_contract_catalog(workbook_path)
    assert catalog["contracts"]["XXA"]["authority"] == "nx_workbook_baseline"
    assert catalog["contracts"]["XXA"]["field_categories"]["学校名称"] == "业务属性字段"
    monkeypatch.setenv("GDA_STANDARD_CONTRACT_XLSX", str(workbook_path))
    mapped = OfflineIngestStore._map_layer(
        {
            "name": "XXA",
            "geometry_type": "Polygon",
            "srid": 4490,
            "fields": [{"name": "OBJECTID"}, {"name": "学校名称"}, {"name": "产生时间"}],
        }
    )
    assert mapped["mapping"]["status"] == "accepted"
    assert mapped["mapping"]["auto_publish"] is True
    assert mapped["contract"]["authority"] == "nx_workbook_baseline"


def test_paper9_sensitive_default_aliases_require_authority_contract():
    mapped = OfflineIngestStore._map_layer(
        {
            "name": "STBHHX",
            "geometry_type": "Polygon",
            "srid": 4490,
            "fields": [{"name": "BSM"}],
        }
    )
    assert mapped["mapping"]["status"] == "manual_review"
    assert mapped["mapping"]["contract_authority"] == "default_alias_candidate"
