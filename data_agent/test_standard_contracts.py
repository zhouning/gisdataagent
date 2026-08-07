from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from data_agent.offline_ingest import OfflineIngestStore
from data_agent.standard_contracts import (
    compile_standard_contract_catalog,
    load_inventory_field_baseline,
    validate_contract_catalog,
)


def _write_json(path: Path, payload: dict) -> Path:
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return path


def test_compile_contract_is_runtime_baseline_without_attribute_authority(tmp_path):
    roles = _write_json(
        tmp_path / "roles.json",
        {
            "version": "2026-test",
            "source_documents": ["standard.docx"],
            "roles": {
                "parcel": {
                    "standard_tables": [
                        {"table_code": "DLTB", "table_alias_zh": "地类图斑", "module": "调查"}
                    ],
                    "geometry_type": "Polygon",
                    "required_fields": ["BSM", "DLBM"],
                    "recommended_fields": ["TBMJ"],
                    "field_rules": {"TBMJ": {"type": "number", "unit": "m2"}},
                    "twm_binding": {"object_id": "BSM"},
                }
            },
        },
    )
    aliases = _write_json(
        tmp_path / "aliases.json",
        {"field_aliases": {"BSM": "标识码", "DLBM": "地类编码", "TBMJ": "图斑面积"}},
    )
    domains = _write_json(tmp_path / "domains.json", {"domains": {}})
    catalog = compile_standard_contract_catalog(
        role_contracts_path=roles,
        field_aliases_path=aliases,
        value_domains_path=domains,
    )
    assert catalog["authority"] == "nx_workbook_baseline"
    assert catalog["runtime_baseline_ready"] is True
    assert (
        catalog["contracts"]["DLTB"]["publication_gate"]
        == "per_dataset_runtime_validation"
    )
    assert catalog["contracts"]["DLTB"]["fields"][0]["primary_key"] is True
    assert catalog["contracts"]["DLTB"]["fields"][2]["data_type"] == "number"
    assert catalog["inventory_resolution"]["counts"]["unresolved"] == 0
    assert validate_contract_catalog(catalog)  # no EA-standard signature


def test_ea_evidence_is_recorded_in_runtime_baseline(tmp_path):
    roles = _write_json(
        tmp_path / "roles.json",
        {
            "roles": {
                "p": {
                    "standard_tables": [
                        {"table_code": "PDT", "table_alias_zh": "坡度图", "module": "底图"}
                    ],
                    "geometry_type": "Polygon",
                    "required_fields": ["BSM"],
                    "recommended_fields": [],
                    "field_rules": {},
                }
            }
        },
    )
    ea = tmp_path / "ea.csv"
    ea.write_text(
        "standard_table,match_mode,ea_object_id,ea_name,ea_fields,field_coverage_pct\n"
        "PDT,logical_title_exact_domain,42,坡度图,5,100\n",
        encoding="utf-8",
    )
    catalog = compile_standard_contract_catalog(
        role_contracts_path=roles, ea_table_comparison_path=ea
    )
    contract = catalog["contracts"]["PDT"]
    assert contract["authority"] == "nx_project_baseline"
    assert contract["ea_evidence"][0]["ea_object_id"] == "42"
    assert validate_contract_catalog(catalog, require_authoritative=False) == []
    assert validate_contract_catalog(catalog, require_authoritative=True)


def test_v2_contract_maps_canonical_codes_and_exposes_type_gap(tmp_path, monkeypatch):
    roles = _write_json(
        tmp_path / "roles.json",
        {
            "roles": {
                "p": {
                    "standard_tables": [{"table_code": "DLTB", "table_alias_zh": "地类图斑"}],
                    "geometry_type": "Polygon",
                    "required_fields": ["BSM"],
                    "recommended_fields": [],
                    "field_rules": {"BSM": {"type": "string"}},
                    "twm_binding": {"object_id": "BSM"},
                }
            }
        },
    )
    aliases = _write_json(tmp_path / "aliases.json", {"field_aliases": {"BSM": "标识码"}})
    catalog_path = tmp_path / "catalog.json"
    catalog_path.write_text(
        json.dumps(
            compile_standard_contract_catalog(
                role_contracts_path=roles, field_aliases_path=aliases
            ),
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("GDA_STANDARD_CONTRACTS", str(catalog_path))
    mapped = OfflineIngestStore._map_layer(
        {
            "name": "DLTB",
            "geometry_type": "Polygon",
            "srid": 4490,
            "fields": [{"name": "BSM", "type": "String"}],
        }
    )
    assert mapped["mapping"]["matched_fields"] == ["BSM"]
    assert mapped["mapping"]["contract_authority"] == "nx_project_baseline"
    assert mapped["mapping"]["status"] == "accepted"
    assert mapped["mapping"]["auto_publish"] is True
    assert mapped["mapping"]["field_mappings"][0]["standard_type"] == "string"


def test_compile_merges_all_workbook_layers_into_runtime_baseline(tmp_path):
    roles = _write_json(tmp_path / "roles.json", {"roles": {}})
    workbook_path = tmp_path / "shp-fields.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "数据表汇总"
    summary.append(
        [
            "序号",
            "图层/数据表代码",
            "中文名称",
            "几何类型",
            "可辨字段数",
            "来源照片",
            "名称来源",
            "完整性/核验说明",
        ]
    )
    summary.append([1, "XXA", "学校（面）", "面", 2, "照片01", "截图显示", "字段完整"])
    details = workbook.create_sheet("字段明细")
    details.append(
        [
            "总序号",
            "图层/数据表代码",
            "中文名称",
            "几何类型",
            "字段序号",
            "字段名称",
            "字段类别",
            "来源照片",
        ]
    )
    details.append([1, "XXA", "学校（面）", "面", 1, "OBJECTID", "系统/几何字段", "照片01"])
    details.append([2, "XXA", "学校（面）", "面", 2, "学校名称", "业务属性字段", "照片01"])
    workbook.save(workbook_path)

    catalog = compile_standard_contract_catalog(
        role_contracts_path=roles,
        shp_workbook=workbook_path,
    )

    contract = catalog["contracts"]["XXA"]
    assert contract["authority"] == "nx_workbook_baseline"
    assert contract["geometry_type"] == "Polygon"
    assert contract["required_fields"] == ["学校名称"]
    assert catalog["coverage"]["workbook_contract_dataset_count"] == 1
    assert catalog["coverage"]["workbook_field_count"] == 2


def test_inventory_workbook_field_sheets_are_compiled(tmp_path):
    workbook_path = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "客户提供数据清单梳理"
    inventory.append(["序号", "数据大类", "数据小类", "数据名称"])
    inventory.append([1, "矢量", "土地", "2024变更调查数据DLTB地类图斑"])

    def add_sheet(name, title, headers):
        sheet = workbook.create_sheet(name)
        sheet.append([title])
        sheet.append(headers)

    add_sheet(
        "2024年土地利用现状调查数据",
        "DLTB",
        ["标识码", "要素代码", "图斑编号", "地类编码", "图斑面积"],
    )
    add_sheet("2025年不动产-ZRZ", "ZRZ", ["实体标识码", "标识码", "幢号"])
    planning = workbook.create_sheet("自治区级国土空间总体规划数据（3条）")
    planning.append(["城镇建设适宜性评价结果（面）CZJSSYXPJJG"])
    planning.append(["行政区划代码", "一张图要素代码", "适宜性等级"])
    planning.append(["农业生产适宜性评价结果（面）NYSCSYXPJJG"])
    planning.append(["行政区划代码", "一张图要素代码", "适宜性等级"])
    planning.append(["生态保护重要性评价结果（面）STBHZYXPJJG"])
    planning.append(["行政区划代码", "一张图要素代码", "适宜性等级"])
    add_sheet(
        "县级国土空间规划（仅治理历史文化保护线 中心城区黄绿蓝紫线规）",
        "ZXCQ",
        ["标识码", "要素代码", "行政区代码", "面积"],
    )
    add_sheet(
        "银川市城市存量成果FWJZ2024",
        "FWJZ",
        ["建筑编码", "基底面积", "建筑高度"],
    )
    add_sheet(
        "银川市城市存量成果SQCPG2025",
        "SQCPG",
        ["空间单元代码", "空间单元名称", "建筑密度"],
    )
    workbook.save(workbook_path)

    baseline = load_inventory_field_baseline(workbook_path)
    assert baseline["dataset_count"] == 8
    assert baseline["contracts"]["DLTB"]["required_fields"] == [
        "标识码",
        "要素代码",
        "图斑编号",
        "地类编码",
        "图斑面积",
    ]
    assert len(baseline["contracts"]["DLTB"]["fields"]) == 5
    assert {
        "CZJSSYXPJJG",
        "NYSCSYXPJJG",
        "STBHZYXPJJG",
    }.issubset(baseline["contracts"])
    assert baseline["contracts"]["FWJZ"]["fields"][0]["name"] == "建筑编码"
    assert baseline["contracts"]["SQCPG"]["fields"][2]["name"] == "建筑密度"
