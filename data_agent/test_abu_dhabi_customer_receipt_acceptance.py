from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_agent.uwm.abu_dhabi_flood.customer_receipt_acceptance import accept_customer_receipt


def _register(path: Path) -> None:
    issues = [
        {
            "issue_id": "SWMM-ENG-001",
            "priority": "P0",
            "title": "管径单位、断面形状和工程断面未确认",
        },
        {
            "issue_id": "SWMM-ENG-006",
            "priority": "P0",
            "title": "权威事件降雨和设计暴雨输入缺失",
        },
    ]
    path.write_text(
        json.dumps(
            {
                "schema": "gwm.abu_dhabi_flood.customer_swmm_engineering_issue_register.v1",
                "issues": issues,
            }
        ),
        encoding="utf-8",
    )


def _workbook(path: Path, file_path: Path, *, complete: bool) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in ("问题总览", "模型影响映射", "填写说明"):
        workbook.create_sheet(name)
    sheet = workbook.create_sheet("客户回执")
    headers = [
        "问题编号",
        "优先级",
        "工程问题",
        "当前状态",
        "影响模型层",
        "当前证据（项目填写）",
        "对模型的影响（项目填写）",
        "客户需提供/确认内容（项目填写）",
        "关闭条件（项目填写）",
        "客户交付状态",
        "客户文件名/私有路径",
        "图层/表/工作表",
        "字段映射或数据字典位置",
        "数据来源单位/责任方",
        "版本或快照编号",
        "有效时间/事件范围",
        "水平坐标系（CRS）",
        "垂直基准",
        "单位",
        "质量标识/缺测说明",
        "许可或交付授权",
        "SHA-256",
        "客户责任人",
        "预计提供日期",
        "客户备注",
        "项目审核状态",
        "项目审核意见/关闭证据",
    ]
    sheet.append([None] * len(headers))
    sheet.append([None] * len(headers))
    sheet.append([None] * len(headers))
    sheet.append(headers)
    digest = hashlib.sha256(file_path.read_bytes()).hexdigest()
    base = [
        "P0",
        "",
        "requires_customer_confirmation",
        "EPA SWMM",
        "",
        "",
        "",
        "",
        "已提供待验收",
        str(file_path),
        "layer",
        "dictionary",
        "DMT",
        "v1",
        "2024 event",
        "EPSG:32640",
        "local datum",
        "mm",
        "quality flags",
        "customer authority",
        digest,
        "owner",
        "2026-08-23",
        "",
        "待审核",
        "",
    ]
    if not complete:
        base[13] = ""
    for issue_id, title in (
        ("SWMM-ENG-001", "管径单位、断面形状和工程断面未确认"),
        ("SWMM-ENG-006", "权威事件降雨和设计暴雨输入缺失"),
    ):
        row = [issue_id, *base]
        row[2] = title
        sheet.append(row)
    workbook.save(path)


def test_empty_customer_receipt_is_fail_closed(tmp_path: Path):
    register = tmp_path / "register.json"
    _register(register)
    data = tmp_path / "data.txt"
    data.write_text("authoritative", encoding="utf-8")
    workbook = tmp_path / "receipt.xlsx"
    _workbook(workbook, data, complete=False)
    payload = accept_customer_receipt(workbook, register, [tmp_path])
    assert payload["status"] == "customer_receipt_requires_action"
    assert payload["summary"]["accepted_issue_count"] == 0
    assert payload["admission"]["gwm_training_admitted"] is False
    assert all("metadata_missing" in item["reasons"] for item in payload["issues"])


def test_matching_hash_and_metadata_are_accepted_but_do_not_open_model_gates(tmp_path: Path):
    register = tmp_path / "register.json"
    _register(register)
    data = tmp_path / "data.txt"
    data.write_text("authoritative", encoding="utf-8")
    workbook = tmp_path / "receipt.xlsx"
    _workbook(workbook, data, complete=True)
    payload = accept_customer_receipt(workbook, register, [tmp_path])
    assert payload["status"] == "customer_receipt_accepted"
    assert payload["summary"]["accepted_issue_count"] == 2
    assert all(item["accepted"] for item in payload["issues"])
    assert payload["admission"]["customer_authoritative_delivery_complete"] is True
    assert payload["admission"]["engineering_calibration_admitted"] is False
    assert payload["admission"]["traditional_model_admitted"] is False
    assert payload["admission"]["gwm_training_admitted"] is False


def test_hash_mismatch_is_returned_for_completion(tmp_path: Path):
    register = tmp_path / "register.json"
    _register(register)
    data = tmp_path / "data.txt"
    data.write_text("changed", encoding="utf-8")
    workbook = tmp_path / "receipt.xlsx"
    _workbook(workbook, data, complete=True)
    loaded = load_workbook(workbook)
    sheet = loaded["客户回执"]
    sheet.cell(5, 22).value = "0" * 64
    loaded.save(workbook)
    payload = accept_customer_receipt(workbook, register, [tmp_path])
    assert payload["summary"]["accepted_issue_count"] == 1
    assert "file_1_sha256_mismatch" in payload["issues"][0]["reasons"]


def test_public_proxy_is_never_accepted_as_customer_authority(tmp_path: Path):
    register = tmp_path / "register.json"
    _register(register)
    data = tmp_path / "data.txt"
    data.write_text("proxy", encoding="utf-8")
    workbook = tmp_path / "receipt.xlsx"
    _workbook(workbook, data, complete=True)
    loaded = load_workbook(workbook)
    sheet = loaded["客户回执"]
    sheet.cell(5, 14).value = "Open-Meteo public proxy"
    loaded.save(workbook)
    payload = accept_customer_receipt(workbook, register, [tmp_path])
    assert payload["issues"][0]["accepted"] is False
    assert "public_proxy_cannot_close_customer_issue" in payload["issues"][0]["reasons"]
