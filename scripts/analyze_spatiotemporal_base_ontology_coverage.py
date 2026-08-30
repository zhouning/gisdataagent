#!/usr/bin/env python3
"""Audit the 0804 attachment catalog against the curated ontology.

The input workbook is read-only. Outputs are deterministic review artifacts
that distinguish source-field presence from actual domain-semantic coverage.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from data_agent.ontology.domain_model import (
    CLASS_SEEDS,
    CURATED_MODEL_VERSION,
    DATA_PROPERTIES,
    OBJECT_PROPERTIES,
)

DEFAULT_INPUT = Path("/Users/zhouning/Downloads/时空基底数据属性挂接0804.xlsx")
DEFAULT_OUTPUT = Path("docs/analysis/natural-resource-ontology-attachment-coverage-0804")
OLD_PACKAGE = Path("data_agent/ontology/packages/natural_resource_one_map/2.1.0")
FIELD_CODE_RE = re.compile(r"([A-Z][A-Z0-9_]{1,})$")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalized(value: str) -> str:
    return "".join(character for character in value.casefold().strip() if character.isalnum())


def _split_label_code(value: str) -> tuple[str, str]:
    match = FIELD_CODE_RE.search(value)
    if not match:
        return value.strip(), ""
    return value[: match.start()].strip(), match.group(1)


def _merged_value_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                values[(row, column)] = top_left
    return values


def _cell(sheet: openpyxl.worksheet.worksheet.Worksheet, merged: dict, row: int, column: int) -> str:
    return _text(merged.get((row, column), sheet.cell(row, column).value))


def _workbook_rows(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        merged = _merged_value_map(sheet)
        if sheet.title == "Sheet1":
            for row_number in range(3, sheet.max_row + 1):
                attribute_raw = _cell(sheet, merged, row_number, 4)
                if not attribute_raw:
                    continue
                attribute_label, attribute_code = _split_label_code(attribute_raw)
                layer_raw = _cell(sheet, merged, row_number, 3)
                layer_label, layer_code = _split_label_code(layer_raw)
                rows.append({
                    "sheet": sheet.title,
                    "excel_row": row_number,
                    "data_category": _cell(sheet, merged, row_number, 2),
                    "entity_code": "",
                    "entity_label": _cell(sheet, merged, row_number, 5),
                    "source_layer": layer_label,
                    "source_layer_code": layer_code,
                    "source_data": _cell(sheet, merged, row_number, 2),
                    "source_attribute": attribute_label,
                    "source_field_code": attribute_code,
                    "source_attribute_raw": attribute_raw,
                })
        elif sheet.title == "Sheet2":
            for row_number in range(4, sheet.max_row + 1):
                attribute_raw = _cell(sheet, merged, row_number, 5)
                if not attribute_raw:
                    continue
                attribute_label, attribute_code = _split_label_code(attribute_raw)
                layer_raw = _cell(sheet, merged, row_number, 6)
                layer_label, layer_code = _split_label_code(layer_raw)
                rows.append({
                    "sheet": sheet.title,
                    "excel_row": row_number,
                    "data_category": _cell(sheet, merged, row_number, 2),
                    "entity_code": _cell(sheet, merged, row_number, 3),
                    "entity_label": _cell(sheet, merged, row_number, 4),
                    "source_layer": layer_label,
                    "source_layer_code": layer_code,
                    "source_data": _cell(sheet, merged, row_number, 7),
                    "source_attribute": attribute_label,
                    "source_field_code": attribute_code,
                    "source_attribute_raw": attribute_raw,
                })
    return rows


def _load_jsonl_gzip(path: Path) -> list[dict[str, Any]]:
    with gzip.open(path, "rt", encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def _old_model_index(package: Path) -> dict[str, set[str]]:
    concepts = _load_jsonl_gzip(package / "concepts.jsonl.gz")
    properties = _load_jsonl_gzip(package / "properties.jsonl.gz")
    return {
        "curated_classes": {
            str(row.get("code"))
            for row in concepts
            if row.get("source_system") == "curated_domain"
        },
        "curated_property_ids": {
            str(row.get("property_id"))
            for row in properties
            if row.get("source_id") == "natural-resource-domain-model-v2"
        },
        "source_codes": {_normalized(str(row.get("code") or "")) for row in properties},
        "source_labels": {_normalized(str(row.get("pref_label") or "")) for row in properties},
    }


def _entity_class(row: dict[str, Any]) -> tuple[str, str]:
    value = " ".join((row["entity_code"], row["entity_label"], row["source_layer"])).strip()
    checks = (
        (("街道办", "村委会"), "AdministrativePlace"),
        (("R_BLD_A", "房屋", "建构筑物"), "Building"),
        (("人工绿地", "公园", "绿地", "公园广场"), "GreenOpenSpace"),
        (("露天体育场", "体育活动"), "SportsFacility"),
        (("福利机构",), "WelfareFacility"),
        (("医疗机构", "医院", "卫生服务"), "MedicalFacility"),
        (("学校", "中小学", "幼儿园"), "EducationalFacility"),
        (("文艺场馆", "文化活动", "科教文卫"), "CulturalFacility"),
        (("公用设施",), "UtilityFacility"),
        (("交通", "停车场", "高速公路服务区"), "TransportStation"),
        (("殡葬",), "Cemetery"),
        (("应急避难",), "EmergencyShelter"),
        (("住宅小区", "Y_LIV_A"), "ResidentialCompound"),
        (("院落", "场院"), "Courtyard"),
        (("H_RIV_A", "地面河流"), "RiverSegment"),
        (("H_CAN_A", "沟渠"), "Canal"),
        (("H_LAK_A", "湖泊池塘"), "SurfaceWaterBody"),
        (("H_RES_A", "水库山塘"), "Reservoir"),
        (("H_WFC_A", "水利附属设施"), "WaterConservancyFacility"),
        (("全部水系",), "WaterSystem"),
        (("地类图斑",), "LandParcel"),
        (("规划用地用海",), "PlannedLandUseArea"),
        (("规划分区",), "PlanningZone"),
        (("村级行政区",), "Village"),
    )
    for terms, class_name in checks:
        if any(term in value for term in terms):
            return class_name, ""
    if "全域实体" in value:
        return "SpatialUnit", "需由具体来源图层细化为地块、规划单元或行政单元。"
    if "城市四线涉及实体" in value:
        return "NaturalResourceEntity", "相交实体类型由实际空间叠置结果确定。"
    return "NaturalResourceEntity", "源实体分类粒度不足，保留自然资源实体上位类。"


def _property_indexes() -> tuple[dict[str, Any], dict[str, list[Any]], dict[str, list[Any]]]:
    by_name = {seed.name: seed for seed in DATA_PROPERTIES}
    by_label: dict[str, list[Any]] = {}
    by_code: dict[str, list[Any]] = {}
    for seed in DATA_PROPERTIES:
        for label in (seed.label, *seed.source_labels):
            by_label.setdefault(_normalized(label), []).append(seed)
        for code in seed.source_codes:
            by_code.setdefault(_normalized(code), []).append(seed)
    return by_name, by_label, by_code


def _resolve_property(
    row: dict[str, Any],
    by_name: dict[str, Any],
    by_label: dict[str, list[Any]],
    by_code: dict[str, list[Any]],
) -> tuple[list[Any], str]:
    label = row["source_attribute"]
    normalized_label = _normalized(label)
    explicit = {
        _normalized("类型"): "facilityType",
        _normalized("足球场地类型"): "footballFieldType",
        _normalized("水域类别名称"): "waterCategoryName",
        _normalized("用途类型"): "reservoirUseType",
        _normalized("容积"): "storageCapacity",
        _normalized("名称"): "displayName",
        _normalized("机构名称"): "displayName",
        _normalized("设施名称"): "displayName",
        _normalized("场馆名称"): "displayName",
        _normalized("学校名称"): "displayName",
        _normalized("医疗机构名称"): "displayName",
        _normalized("应急避难场所名称"): "displayName",
    }
    if normalized_label == _normalized("公共设施建筑占比特殊建筑占比"):
        return [
            by_name["publicFacilityBuildingShare"],
            by_name["specialUseBuildingShare"],
        ], "源单元格疑似粘连两个指标，需拆列并确认各自数值。"
    if normalized_label in explicit:
        return [by_name[explicit[normalized_label]]], ""
    label_matches = by_label.get(normalized_label, [])
    if len(label_matches) == 1:
        return label_matches, ""
    code_matches = by_code.get(_normalized(row["source_field_code"]), [])
    if len(code_matches) == 1:
        return code_matches, ""
    if label_matches:
        return label_matches, "同名源属性需要结合实体类型确认领域属性。"
    if code_matches:
        return code_matches, "源字段代码在多个语义上下文中复用，已按来源图层约束候选。"
    return [], "未找到稳定领域属性。"


def _relation(entity_class: str, property_owner: str, attribute_label: str) -> str:
    if property_owner in {"UrbanFormAssessment", "BuiltEnvironmentAssessment"}:
        return f"{entity_class} hasAssessment {property_owner}"
    if property_owner in {"ServiceCoverageObservation", "WaterBodyObservation"}:
        return f"{entity_class} hasObservation {property_owner}"
    if property_owner == "NaturalResourceRight":
        return f"{entity_class} hasRight NaturalResourceRight"
    if property_owner in {"CadastralParcel", "PlannedLandUseArea", "PlanningZone"} and entity_class != property_owner:
        return f"{entity_class} overlapsSpatialUnit {property_owner}"
    if property_owner == "SurfaceWaterBody" and entity_class == "WaterSystem":
        return "WaterSystem hasWaterFeature SurfaceWaterBody"
    return "direct datatype property"


def _four_line(attribute: str) -> str:
    return {
        "中心城区城市蓝线": "UrbanBlueLine",
        "中心城区城市绿线": "UrbanGreenLine",
        "中心城区城市黄线": "UrbanYellowLine",
        "中心城区城市紫线": "UrbanPurpleLine",
    }.get(attribute, "")


def _coverage_rows(source_rows: list[dict[str, Any]], old: dict[str, set[str]]) -> list[dict[str, Any]]:
    class_names = {seed.name for seed in CLASS_SEEDS}
    property_names, properties_by_label, properties_by_code = _property_indexes()
    results: list[dict[str, Any]] = []
    for sequence, source in enumerate(source_rows, start=1):
        entity_class, entity_note = _entity_class(source)
        line_class = _four_line(source["source_attribute"])
        properties, property_note = _resolve_property(
            source,
            property_names,
            properties_by_label,
            properties_by_code,
        )
        data_quality_pending = source["source_attribute"] == "公共设施建筑占比特殊建筑占比"
        if line_class:
            target_class = line_class
            target_properties = ""
            target_relation = f"{entity_class} intersectsBoundary {line_class}"
        elif properties:
            owners = sorted({seed.owner for seed in properties})
            target_class = "; ".join(owners)
            target_properties = "; ".join(seed.name for seed in properties)
            target_relation = "; ".join(
                sorted({_relation(entity_class, owner, source["source_attribute"]) for owner in owners})
            )
        else:
            target_class = entity_class
            target_properties = ""
            target_relation = ""

        normalized_code = _normalized(source["source_field_code"])
        normalized_label = _normalized(source["source_attribute"])
        source_metadata_present = bool(
            (normalized_code and normalized_code in old["source_codes"])
            or normalized_label in old["source_labels"]
        )
        target_class_names = {item.strip() for item in target_class.split(";") if item.strip()}
        before_classes_exist = bool(target_class_names) and target_class_names.issubset(old["curated_classes"])
        before_properties_exist = bool(properties) and all(
            f"gda:nr:property:{seed.name}" in old["curated_property_ids"] for seed in properties
        )
        if data_quality_pending:
            before_status = "数据质量待确认"
        elif before_classes_exist and (before_properties_exist or line_class):
            before_status = "已有领域语义"
        elif source_metadata_present:
            before_status = "仅有源元数据"
        elif before_classes_exist:
            before_status = "可映射但需补充"
        else:
            before_status = "完全缺失"

        if data_quality_pending:
            after_status = "数据质量待确认"
        elif line_class and line_class in class_names:
            after_status = "类与关系已覆盖"
        elif properties and all(seed.owner in class_names for seed in properties):
            after_status = "类与属性已覆盖"
        else:
            after_status = "未覆盖"

        notes = [note for note in (entity_note, property_note) if note]
        if source["source_field_code"] in {"CNN", "TYPE"}:
            notes.append("CNN/TYPE 仅是源字段代码，按来源实体上下文映射，不作为两个领域属性。")
        results.append({
            "sequence": sequence,
            **source,
            "attachment_entity_class": entity_class,
            "target_domain_class": target_class,
            "target_data_property": target_properties,
            "target_object_relation": target_relation,
            "v2_1_source_metadata_present": source_metadata_present,
            "v2_1_coverage_status": before_status,
            "v2_2_coverage_status": after_status,
            "model_version": CURATED_MODEL_VERSION,
            "modeling_note": " ".join(notes),
        })
    return results


QUALITY_ISSUES = [
    {
        "severity": "blocking_for_value_load",
        "issue": "公共设施建筑占比特殊建筑占比疑似两个指标粘连",
        "action": "拆为公共设施建筑占比和特殊建筑占比，分别确认数值、单位和计算公式。",
    },
    {
        "severity": "mapping_rule_required",
        "issue": "CNN 与 TYPE 在多个图层复用为类型字段",
        "action": "以来源图层和目标实体类作为映射上下文，禁止仅凭字段代码自动合并。",
    },
    {
        "severity": "mapping_rule_required",
        "issue": "R_GRO_A 和 Y_PSE_A 被复用于不同实体语义",
        "action": "以业务子类和源图层联合键区分体育场、人工绿地、科教文卫、公用设施和公园广场。",
    },
    {
        "severity": "definition_required",
        "issue": "无代码指标缺少单位、值域、统计周期和计算方法",
        "action": "指标值必须关联 Measurement、MeasurementUnit、AggregationContext 和证据来源。",
    },
    {
        "severity": "spatial_join_required",
        "issue": "院落挂接宗地、规划单元和评估指标依赖空间关联",
        "action": "保留源对象及空间叠置关系，不把宗地边界、规划分类或评价值误当成院落固有属性。",
    },
    {
        "severity": "identity_required",
        "issue": "街道办、村委会地点与行政范围被合并描述",
        "action": "分别建立机构、地点和行政范围，并通过 officeOf 与 representsAdministrativeUnit 连接。",
    },
]


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    matrix = workbook.active
    matrix.title = "逐字段覆盖矩阵"
    headers = list(rows[0])
    matrix.append(headers)
    for row in rows:
        matrix.append([row.get(header) for header in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in matrix[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    matrix.freeze_panes = "A2"
    matrix.auto_filter.ref = matrix.dimensions
    for column in matrix.columns:
        letter = column[0].column_letter
        width = min(48, max(10, max(len(_text(cell.value)) for cell in column) + 2))
        matrix.column_dimensions[letter].width = width
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_sheet = workbook.create_sheet("汇总")
    summary_sheet.append(["项目", "结果"])
    for key, value in summary.items():
        summary_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    quality_sheet = workbook.create_sheet("数据质量问题")
    quality_sheet.append(["严重度", "问题", "处置建议"])
    for issue in QUALITY_ISSUES:
        quality_sheet.append([issue["severity"], issue["issue"], issue["action"]])
    for sheet in (summary_sheet, quality_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                80, max(12, max(len(_text(cell.value)) for cell in column) + 2)
            )
            for cell in column:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def _write_markdown(path: Path, summary: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    before = summary["v2_1_coverage_status_counts"]
    after = summary["v2_2_coverage_status_counts"]
    lines = [
        "# 时空基底数据属性挂接 0804 自然资源本体覆盖分析",
        "",
        f"- 输入文件 SHA-256：`{summary['input_sha256']}`",
        f"- 清单属性行：`{summary['attribute_row_count']}`（Sheet1 `{summary['sheet_row_counts']['Sheet1']}`，Sheet2 `{summary['sheet_row_counts']['Sheet2']}`）",
        f"- 扩展模型版本：`{CURATED_MODEL_VERSION}`",
        f"- 策划类 / 对象关系 / 领域数据属性：`{summary['curated_class_count']}` / `{summary['object_property_count']}` / `{summary['data_property_count']}`",
        "",
        "## 结论",
        "",
        "V2.1.0 不能完整覆盖该清单。它已经具备土地、宗地、权利、规划控制和河流/湖泊/水库骨架，但大量行只有来源结构字段，缺少建成环境、水系细分、行政地点、指标观测和稳定领域数据属性。",
        "",
        f"按 V2.2.0 扩展后，`{after.get('类与属性已覆盖', 0) + after.get('类与关系已覆盖', 0)}` / `{len(rows)}` 行具有明确的领域类、属性或关系落点；剩余 `{after.get('数据质量待确认', 0)}` 行是同一单元格粘连两个指标，必须先修复源清单，不能猜测装载。",
        "",
        "## 覆盖变化",
        "",
        "| 阶段 | 状态 | 行数 |",
        "| --- | --- | ---: |",
    ]
    for status, count in sorted(before.items()):
        lines.append(f"| V2.1.0 | {status} | {count} |")
    for status, count in sorted(after.items()):
        lines.append(f"| V2.2.0 | {status} | {count} |")
    lines.extend([
        "",
        "## 新增语义层",
        "",
        "1. 建成环境：建构筑物、房屋、住宅小区、院落、教育/医疗/文化/福利/体育/公用/交通/避难/绿地/殡葬设施。",
        "2. 水系：水系整体、河段、沟渠、池塘、水库和水利附属设施，并区分整体水系与组成水体。",
        "3. 行政地点：街道、村、社区、行政机构、机构地点和行政范围分别建模。",
        "4. 规划管控：规划用地用海单元、规划分区、城市四线及其与实体的空间相交关系。",
        "5. 指标观测：城市形态评价、建成环境评价、公共服务覆盖观测、测量值、单位、聚合上下文、出行方式、阈值和人口分母。",
        "6. 数据属性：建筑、权属、规划、住房、评价、可达性和水文属性均采用稳定英文 IRI，并记录原字段代码和中文名称。",
        "",
        "## 关键建模原则",
        "",
        "- `Land` 的直接子类仍只有 `AgriculturalLand`、`ConstructionLand`、`UnusedLand`。",
        "- `LandParcel` 仍属于 `SpatialUnit`，通过 `spatiallyRepresents` 表征 `Land`。",
        "- 宗地四至、规划分类和评价指标通过权利、空间叠置或观测关系挂接，不降格为院落的固有字段。",
        "- CNN/TYPE 是源字段代码；领域语义由目标类和来源图层共同确定。",
        "- 可达覆盖率必须保留出行方式、时间阈值、统计空间、人口口径、时点和计算证据。",
        "",
        "## 数据质量问题",
        "",
    ])
    for issue in QUALITY_ISSUES:
        lines.append(f"- **{issue['issue']}**：{issue['action']}")
    lines.extend([
        "",
        "## 审查产物",
        "",
        "- `natural-resource-ontology-attachment-coverage-0804.csv`：逐字段覆盖矩阵。",
        "- `natural-resource-ontology-attachment-coverage-0804.json`：机器可读矩阵和汇总。",
        "- `natural-resource-ontology-attachment-coverage-0804.xlsx`：适合业务审阅和筛选的矩阵。",
        "",
        "原始工作簿保持只读，未覆盖、改名或写回。",
        "",
    ])
    path.write_text("\n".join(lines), encoding="utf-8")


def build(input_path: Path, output_prefix: Path, old_package: Path) -> dict[str, Any]:
    source_rows = _workbook_rows(input_path)
    sheet_counts = Counter(row["sheet"] for row in source_rows)
    if sheet_counts != {"Sheet1": 129, "Sheet2": 140}:
        raise RuntimeError(f"unexpected workbook row counts: {dict(sheet_counts)}")
    rows = _coverage_rows(source_rows, _old_model_index(old_package))
    before_counts = Counter(row["v2_1_coverage_status"] for row in rows)
    after_counts = Counter(row["v2_2_coverage_status"] for row in rows)
    summary = {
        "input_path": str(input_path),
        "input_sha256": hashlib.sha256(input_path.read_bytes()).hexdigest(),
        "attribute_row_count": len(rows),
        "sheet_row_counts": dict(sheet_counts),
        "unique_attribute_label_count": len({row["source_attribute"] for row in rows}),
        "unique_source_field_code_count": len({row["source_field_code"] for row in rows if row["source_field_code"]}),
        "unique_source_layer_count": len({(row["source_layer"], row["source_layer_code"]) for row in rows}),
        "v2_1_coverage_status_counts": dict(before_counts),
        "v2_2_coverage_status_counts": dict(after_counts),
        "curated_model_version": CURATED_MODEL_VERSION,
        "curated_class_count": len(CLASS_SEEDS),
        "object_property_count": len(OBJECT_PROPERTIES),
        "data_property_count": len(DATA_PROPERTIES),
        "quality_issue_count": len(QUALITY_ISSUES),
    }
    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    _write_csv(output_prefix.with_suffix(".csv"), rows)
    output_prefix.with_suffix(".json").write_text(
        json.dumps(
            {"summary": summary, "quality_issues": QUALITY_ISSUES, "coverage": rows},
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    _write_xlsx(output_prefix.with_suffix(".xlsx"), rows, summary)
    _write_markdown(output_prefix.with_suffix(".md"), summary, rows)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-prefix", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--old-package", type=Path, default=OLD_PACKAGE)
    args = parser.parse_args()
    print(json.dumps(build(args.input, args.output_prefix, args.old_package), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
