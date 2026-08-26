#!/usr/bin/env python3
"""Run six 180-minute Abu Dhabi Zone B DDF scenarios on the full SWMM network."""

from __future__ import annotations

import argparse
import csv
import json
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from data_agent.abu_dhabi_flood_scenario_service import (
    DEFAULT_RUN_ROOT,
    _parse_node_hydraulic_results,
    _run_worker,
    public_run,
    validate_scenario,
)
from data_agent.abu_dhabi_zone_b_design_storm import (
    SUPPORTED_RETURN_PERIODS,
    ZONE_B_DDF_DEPTH_MM,
    ZONE_B_DURATIONS_MINUTES,
    ZONE_B_EVIDENCE_SHA256,
    ZONE_B_IDF_INTENSITY_MM_PER_HOUR,
    ZONE_B_SOURCE_YEAR,
    ZONE_B_TABLE_TITLE,
    zone_b_180_minute_hyetograph,
)


BATCH_SCHEMA = "gwm.abu_dhabi_flood.zone_b_design_storm_batch.v1"
DEFAULT_BATCH_PREFIX = "abu-zone-b-ddf-180m"


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _scenario(return_period: int, *, tail_minutes: int, report_step_minutes: int) -> dict[str, Any]:
    depth = ZONE_B_DDF_DEPTH_MM[return_period][6]
    return validate_scenario(
        {
            "scope": "citywide",
            "rainfallMode": "design_storm",
            "startTime": "2022-01-01T00:00",
            "durationMinutes": 180,
            "tailMinutes": tail_minutes,
            "totalDepthMm": depth,
            "rainfallPattern": "official_zone_b_ddf_abm",
            "returnPeriodYears": return_period,
            "peakPosition": 40,
            "spatialPattern": "uniform",
            "pipeScope": "none",
            "blockagePercent": 0,
            "pipeCapacityMultiplier": 1,
            "pumpEnabled": False,
            "pumpCapacityMultiplier": 1,
            "outfallMode": "open",
            "outfallLevelM": 0,
            "outputIntervalMinutes": report_step_minutes,
        }
    )


def _run_one(
    run_id: str,
    return_period: int,
    *,
    tail_minutes: int,
    report_step_minutes: int,
) -> dict[str, Any]:
    _run_worker(
        run_id,
        _scenario(
            return_period,
            tail_minutes=tail_minutes,
            report_step_minutes=report_step_minutes,
        ),
    )
    return public_run(run_id)


def _node_summary(report_path: Path) -> dict[str, Any]:
    nodes = _parse_node_hydraulic_results(report_path)
    depths = [float(row.get("max_water_depth_m", 0.0) or 0.0) for row in nodes.values()]
    overflows = [float(row.get("max_overflow_or_flooding_m3s", 0.0) or 0.0) for row in nodes.values()]
    return {
        "node_result_count": len(nodes),
        "nodes_depth_ge_0_05_m": sum(value >= 0.05 for value in depths),
        "nodes_depth_ge_0_15_m": sum(value >= 0.15 for value in depths),
        "nodes_depth_ge_0_30_m": sum(value >= 0.30 for value in depths),
        "nodes_depth_ge_0_50_m": sum(value >= 0.50 for value in depths),
        "nodes_depth_ge_1_00_m": sum(value >= 1.00 for value in depths),
        "nodes_with_overflow": sum(value > 0.0 for value in overflows),
        "maximum_node_depth_m": max(depths, default=0.0),
        "maximum_node_overflow_m3s": max(overflows, default=0.0),
        "summed_node_flood_volume_million_litres": sum(
            float(row.get("total_flood_volume_million_litres", 0.0) or 0.0)
            for row in nodes.values()
        ),
    }


def _summarize_run(run_root: Path, return_period: int, run: dict[str, Any]) -> dict[str, Any]:
    partition = (run.get("partitions") or [{}])[0]
    native_root = run_root / str(run["run_id"]) / "full_city" / "native_swmm_results"
    report_paths = sorted(native_root.glob("*.rpt"))
    output_paths = sorted(native_root.glob("*.out"))
    receipt_path = run_root / str(run["run_id"]) / "full_city" / "swmm_execution_receipt.json"
    receipt = json.loads(receipt_path.read_text(encoding="utf-8")) if receipt_path.is_file() else {}
    node_summary = _node_summary(report_paths[0]) if report_paths else {}
    return {
        "return_period_years": return_period,
        "published_180_minute_mean_intensity_mm_per_hour": ZONE_B_IDF_INTENSITY_MM_PER_HOUR[return_period][6],
        "published_180_minute_depth_mm": ZONE_B_DDF_DEPTH_MM[return_period][6],
        "status": run.get("status"),
        "run_id": run.get("run_id"),
        "started_at": run.get("started_at"),
        "finished_at": run.get("finished_at"),
        "rainfall_stats": (run.get("scenario") or {}).get("rainfall_stats"),
        "hydraulic_summary": partition.get("result_summary"),
        "node_summary": node_summary,
        "strict_quality_gates": receipt.get("strict_quality_gates"),
        "artifacts": {
            "scenario_manifest": str(run_root / str(run["run_id"]) / "scenario_manifest.json"),
            "scenario_input": str(run_root / str(run["run_id"]) / "full_city" / "scenario.inp"),
            "execution_receipt": str(receipt_path),
            "native_report": str(report_paths[0]) if report_paths else None,
            "native_binary_output": str(output_paths[0]) if output_paths else None,
        },
    }


def _write_comparison_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = [
        "return_period_years",
        "published_180_minute_mean_intensity_mm_per_hour",
        "published_180_minute_depth_mm",
        "status",
        "run_id",
        "external_outflow_million_litres",
        "flooding_loss_million_litres",
        "runoff_continuity_error_percent",
        "routing_continuity_error_percent",
        "strict_quality_passed",
        "node_result_count",
        "nodes_depth_ge_0_05_m",
        "nodes_depth_ge_0_15_m",
        "nodes_depth_ge_0_30_m",
        "nodes_depth_ge_0_50_m",
        "nodes_depth_ge_1_00_m",
        "nodes_with_overflow",
        "maximum_node_depth_m",
        "maximum_node_overflow_m3s",
        "summed_node_flood_volume_million_litres",
        "native_report",
        "native_binary_output",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            hydraulic = row.get("hydraulic_summary") or {}
            nodes = row.get("node_summary") or {}
            artifacts = row.get("artifacts") or {}
            writer.writerow(
                {
                    **{key: row.get(key) for key in fields},
                    **{key: hydraulic.get(key) for key in (
                        "external_outflow_million_litres",
                        "flooding_loss_million_litres",
                        "runoff_continuity_error_percent",
                        "routing_continuity_error_percent",
                    )},
                    "strict_quality_passed": (row.get("strict_quality_gates") or {}).get("passed"),
                    **{key: nodes.get(key) for key in fields if key in nodes},
                    "native_report": artifacts.get("native_report"),
                    "native_binary_output": artifacts.get("native_binary_output"),
                }
            )


def _write_hyetograph_csv(path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["return_period_years", "interval_start_minutes", "interval_end_minutes", "intensity_mm_per_hour", "interval_depth_mm"])
        for return_period in SUPPORTED_RETURN_PERIODS:
            values, _ = zone_b_180_minute_hyetograph(
                return_period,
                start=datetime(2022, 1, 1),
                peak_position_percent=40,
                tail_minutes=180,
            )
            for index, (_, intensity) in enumerate(values[:36]):
                writer.writerow([return_period, index * 5, (index + 1) * 5, f"{intensity:.8f}", f"{intensity / 12.0:.8f}"])


def _write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    tail_minutes: int,
    report_step_minutes: int,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SWMM comparison"
    headers = [
        "Return period (years)", "180-min intensity (mm/h)", "180-min depth (mm)", "Run status",
        "Flooding loss (million L)", "External outflow (million L)", "Nodes depth >=0.05m",
        "Nodes depth >=0.30m", "Nodes depth >=0.50m", "Nodes with overflow", "Maximum node depth (m)",
        "Routing continuity error (%)", "Strict quality passed", "Run ID",
    ]
    sheet.append(headers)
    for row in rows:
        hydraulic = row.get("hydraulic_summary") or {}
        nodes = row.get("node_summary") or {}
        sheet.append([
            row["return_period_years"], row["published_180_minute_mean_intensity_mm_per_hour"],
            row["published_180_minute_depth_mm"], row.get("status"), hydraulic.get("flooding_loss_million_litres"),
            hydraulic.get("external_outflow_million_litres"), nodes.get("nodes_depth_ge_0_05_m"),
            nodes.get("nodes_depth_ge_0_30_m"), nodes.get("nodes_depth_ge_0_50_m"), nodes.get("nodes_with_overflow"),
            nodes.get("maximum_node_depth_m"), hydraulic.get("routing_continuity_error_percent"),
            (row.get("strict_quality_gates") or {}).get("passed"), row.get("run_id"),
        ])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width

    inputs = workbook.create_sheet("Official inputs")
    inputs.append(["Duration (min)", *[f"{rp}-year depth (mm)" for rp in SUPPORTED_RETURN_PERIODS]])
    for index, duration in enumerate(ZONE_B_DURATIONS_MINUTES):
        inputs.append([duration, *[ZONE_B_DDF_DEPTH_MM[rp][index] for rp in SUPPORTED_RETURN_PERIODS]])
    assumptions = workbook.create_sheet("Scope and assumptions")
    assumptions.append(["Item", "Value"])
    assumptions_rows = [
        ("Source year", ZONE_B_SOURCE_YEAR),
        ("Source table", ZONE_B_TABLE_TITLE),
        ("Evidence SHA-256", ZONE_B_EVIDENCE_SHA256),
        ("Rain duration", "180 minutes"),
        ("Recession tail", f"{tail_minutes} minutes"),
        ("Result timeline interval", f"{report_step_minutes} minutes"),
        ("Temporal distribution", "Alternating block, 5-minute interval"),
        ("Interpolation", "Log-duration/log-depth between published DDF points"),
        ("Peak position", "40%; modeling assumption pending official temporal pattern"),
        ("Spatial distribution", "Uniform Zone B; pending authoritative zone geometry"),
        ("Model status", "Diagnostic, uncalibrated, not engineering admitted"),
    ]
    for item in assumptions_rows:
        assumptions.append(item)
    workbook.save(path)


def run_batch(
    *,
    run_root: Path,
    batch_id: str,
    workers: int,
    tail_minutes: int,
    report_step_minutes: int,
    timeout_seconds: int,
) -> dict[str, Any]:
    run_root = run_root.expanduser().resolve()
    run_root.mkdir(parents=True, exist_ok=True)
    os.environ["ABU_DHABI_SWMM_INTERACTIVE_RUN_ROOT"] = str(run_root)
    os.environ["ABU_DHABI_SWMM_TIMEOUT_SECONDS"] = str(timeout_seconds)
    batch_root = run_root / "batches" / batch_id
    batch_root.mkdir(parents=True, exist_ok=True)
    _write_hyetograph_csv(batch_root / "zone_b_180_minute_hyetographs.csv")
    started_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    results_by_period: dict[int, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="zone-b-swmm") as executor:
        futures = {
            executor.submit(
                _run_one,
                f"{batch_id}-rp{return_period:03d}",
                return_period,
                tail_minutes=tail_minutes,
                report_step_minutes=report_step_minutes,
            ): return_period
            for return_period in SUPPORTED_RETURN_PERIODS
        }
        for future in as_completed(futures):
            return_period = futures[future]
            results_by_period[return_period] = future.result()
            print(f"completed return_period={return_period} status={results_by_period[return_period].get('status')}", flush=True)
    rows = [
        _summarize_run(run_root, return_period, results_by_period[return_period])
        for return_period in SUPPORTED_RETURN_PERIODS
    ]
    manifest = {
        "schema": BATCH_SCHEMA,
        "batch_id": batch_id,
        "status": "completed_with_quality_warnings" if any(row.get("status") != "completed" for row in rows) else "completed",
        "started_at": started_at,
        "finished_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": {
            "year": ZONE_B_SOURCE_YEAR,
            "table": ZONE_B_TABLE_TITLE,
            "evidence_sha256": ZONE_B_EVIDENCE_SHA256,
            "document_title_and_page": "pending_user_confirmation",
        },
        "fixed_scenario": {
            "rain_duration_minutes": 180,
            "recession_tail_minutes": tail_minutes,
            "forcing_step_minutes": 5,
            "report_step_minutes": report_step_minutes,
            "solver_timeout_seconds": timeout_seconds,
            "peak_position_percent": 40,
            "network": "one topology-preserving full-city SWMM diagnostic input",
            "pipe_action": "none",
            "outfall_boundary": "free diagnostic assumption",
        },
        "runs": rows,
        "claim_boundary": [
            "official Zone B DDF depths are used as rainfall-volume inputs",
            "the five-minute alternating-block distribution and 40 percent peak position are assumptions",
            "the customer network model remains uncalibrated and fails strict numerical quality admission",
            "results are diagnostic prototype outputs and are not engineering or city-scale prediction claims",
        ],
    }
    _write_json(batch_root / "batch_manifest.json", manifest)
    _write_comparison_csv(batch_root / "swmm_return_period_comparison.csv", rows)
    _write_workbook(
        batch_root / "swmm_return_period_comparison.xlsx",
        rows,
        tail_minutes=tail_minutes,
        report_step_minutes=report_step_minutes,
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-root", type=Path, default=DEFAULT_RUN_ROOT)
    parser.add_argument("--batch-id", default=f"{DEFAULT_BATCH_PREFIX}-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}")
    parser.add_argument("--workers", type=int, choices=(1, 2, 3), default=2)
    parser.add_argument("--tail-minutes", type=int, choices=range(0, 241, 5), default=60)
    parser.add_argument("--report-step-minutes", type=int, choices=(5, 15, 30, 60), default=30)
    parser.add_argument("--timeout-seconds", type=int, default=2700)
    args = parser.parse_args()
    result = run_batch(
        run_root=args.run_root,
        batch_id=args.batch_id,
        workers=args.workers,
        tail_minutes=args.tail_minutes,
        report_step_minutes=args.report_step_minutes,
        timeout_seconds=args.timeout_seconds,
    )
    print(json.dumps({"batch_id": result["batch_id"], "status": result["status"], "run_count": len(result["runs"])}, indent=2))


if __name__ == "__main__":
    main()
