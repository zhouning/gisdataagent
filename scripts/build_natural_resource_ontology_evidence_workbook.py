#!/usr/bin/env python3
"""Build the V2.3 ontology evidence review workbook."""

from __future__ import annotations

import argparse
import gzip
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_COVERAGE = Path("docs/analysis/natural-resource-ontology-attachment-coverage-0804.json")
DEFAULT_AUDIT = Path("docs/analysis/natural-resource-ontology-evidence-expansion-0804.json")
DEFAULT_PACKAGE = Path("data_agent/ontology/packages/natural_resource_one_map/2.3.0")
DEFAULT_OUTPUT = Path("docs/analysis/natural-resource-ontology-v2.3.0-evidence-review.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="E8F0EC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2DD"))


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    with gzip.open(path, "rt", encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def _flat(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    columns: list[tuple[str, str]],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append([label for _, label in columns])
    for row in rows:
        sheet.append([_flat(row.get(key, "")) for key, _ in columns])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    sheet.auto_filter.ref = sheet.dimensions
    if rows:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for index, (_, label) in enumerate(columns, start=1):
        values = [str(row.get(columns[index - 1][0], "")) for row in rows[:500]]
        width = min(max([len(label), *(len(value) for value in values)] or [len(label)]) + 2, 48)
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(width, 12)


def build(
    coverage_path: Path,
    audit_path: Path,
    package: Path,
    output: Path,
) -> None:
    coverage = _read_json(coverage_path)
    audit = _read_json(audit_path)
    manifest = _read_json(package / "manifest.json")
    active = _read_json(package.parent / "active.json")
    concepts = _read_jsonl(package / "concepts.jsonl.gz")
    properties = _read_jsonl(package / "properties.jsonl.gz")
    relations = _read_jsonl(package / "relations.jsonl.gz")

    curated_source = "natural-resource-domain-model-v2"
    domain_kinds = {
        "DomainClass",
        "InformationClass",
        "ObservationClass",
        "ProcessClass",
        "RoleClass",
        "StateClass",
    }
    domain_classes = [
        row
        for row in concepts
        if row.get("source_id") == curated_source and row.get("kind") in domain_kinds
    ]
    class_by_id = {row["concept_id"]: row for row in domain_classes}
    parent_by_id = {
        row["source_concept_id"]: row["target_concept_id"]
        for row in relations
        if row.get("source_id") == curated_source and row.get("relation_type") == "subClassOf"
    }
    class_rows = []
    for row in domain_classes:
        provenance = row.get("provenance") or {}
        parent = class_by_id.get(parent_by_id.get(row["concept_id"]), {})
        class_rows.append(
            {
                "code": row.get("code"),
                "label": row.get("pref_label"),
                "kind": row.get("kind"),
                "parent_code": parent.get("code"),
                "parent_label": parent.get("pref_label"),
                "domain_id": row.get("domain_id"),
                "definition": row.get("definition"),
                "aliases": row.get("alt_labels"),
                "evidence_status": provenance.get("evidence_status"),
                "evidence_count": len(provenance.get("source_evidence") or []),
            }
        )

    domain_properties = [row for row in properties if row.get("source_id") == curated_source]
    property_rows = []
    for row in domain_properties:
        owner = class_by_id.get(row["owner_concept_id"], {})
        provenance = row.get("provenance") or {}
        property_rows.append(
            {
                "code": row.get("code"),
                "label": row.get("pref_label"),
                "owner_code": owner.get("code"),
                "owner_label": owner.get("pref_label"),
                "datatype": row.get("datatype"),
                "min_count": row.get("min_count"),
                "max_count": row.get("max_count"),
                "source_codes": provenance.get("source_field_codes"),
                "source_labels": provenance.get("source_field_labels"),
                "definition": provenance.get("definition"),
            }
        )

    object_relations = [
        row
        for row in relations
        if row.get("source_id") == curated_source and row.get("relation_type") == "objectProperty"
    ]
    relation_rows = []
    for row in object_relations:
        provenance = row.get("provenance") or {}
        domain = class_by_id.get(row["source_concept_id"], {})
        range_class = class_by_id.get(row["target_concept_id"], {})
        relation_rows.append(
            {
                "name": provenance.get("property_name"),
                "label": row.get("pref_label"),
                "domain_code": domain.get("code"),
                "domain_label": domain.get("pref_label"),
                "range_code": range_class.get("code"),
                "range_label": range_class.get("pref_label"),
                "inverse": provenance.get("inverse_property"),
                "functional": provenance.get("functional"),
                "restriction": provenance.get("restriction"),
            }
        )

    coverage_last_row = len(coverage["coverage"]) + 1
    layer_last_row = len(audit["layers"]) + 1
    field_last_row = len(audit["expanded_fields"]) + 1
    gap_last_row = sum(row["status"] != "matched" for row in audit["layers"]) + 1
    class_last_row = len(class_rows) + 1
    property_last_row = len(property_rows) + 1
    relation_last_row = len(relation_rows) + 1

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("结论摘要")
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 42
    summary.column_dimensions["D"].width = 52
    summary.append(["指标", "值", "判定口径", "来源"])
    summary_rows = [
        (
            "清单最低属性行",
            f"=COUNTA('清单逐项覆盖'!A2:A{coverage_last_row})",
            "必须等于 269",
            "0804 附件逐行解析",
        ),
        (
            "清单图层",
            f"=COUNTA('图层证据匹配'!A2:A{layer_last_row})",
            "必须等于 25",
            "按来源图层与代码分组",
        ),
        (
            "命中 EA/标准结构图层",
            f"=COUNTIF('图层证据匹配'!J2:J{layer_last_row},\"matched\")",
            "有完整结构证据",
            "冻结 2.2.0 来源包",
        ),
        (
            "显式证据缺口",
            f"=COUNTA('证据缺口'!H2:H{gap_last_row})",
            "当前为 5，不得伪造字段",
            "EA 与十册标准无直接结构命中",
        ),
        (
            "EA/标准展开完整字段",
            f"=COUNTA('完整字段语义处置'!A2:A{field_last_row})",
            "必须等于 390",
            "20 个命中结构的全部字段",
        ),
        (
            "未映射或歧义字段",
            f"=COUNTIF('完整字段语义处置'!Q2:Q{field_last_row},\"unresolved_domain_field\")+COUNTIF('完整字段语义处置'!Q2:Q{field_last_row},\"ambiguous_property_mapping\")",
            "必须等于 0",
            "V2.3 字段语义处置门",
        ),
        (
            "领域属性映射字段",
            f"=COUNTIF('完整字段语义处置'!Q2:Q{field_last_row},\"mapped_domain_property\")+COUNTIF('完整字段语义处置'!Q2:Q{field_last_row},\"mapped_related_semantics\")",
            "领域属性及条件性相关语义",
            "V2.3 显式映射",
        ),
        (
            "对象关系映射字段",
            f"=COUNTIF('完整字段语义处置'!Q2:Q{field_last_row},\"mapped_object_relation\")",
            "外键语义不得退化为字符串",
            "V2.3 显式映射",
        ),
        ("领域类", f"=COUNTA('领域类'!A2:A{class_last_row})", "仅现实领域概念", "V2.3 策划模型"),
        (
            "领域数据属性",
            f"=COUNTA('领域属性'!A2:A{property_last_row})",
            "不含来源字段",
            "V2.3 策划模型",
        ),
        (
            "领域对象关系",
            f"=COUNTA('对象关系'!A2:A{relation_last_row})",
            "不含表关联",
            "V2.3 策划模型",
        ),
        (
            "能力问题",
            f"{manifest['stats']['competency_question_passed_count']}/{manifest['stats']['competency_question_count']}",
            "必须全部通过",
            "2.3.0 manifest.json",
        ),
        (
            "SHACL",
            "通过" if manifest["validation_summary"]["shacl_conforms"] else "失败",
            "必须通过",
            "2.3.0 validation-report.json",
        ),
        (
            "OWL-RL 不可满足命名类",
            manifest["stats"]["unsatisfiable_named_class_count"],
            "必须等于 0",
            "2.3.0 semantic-quality-report.json",
        ),
        ("当前线上激活版本", active["semantic_version"], "本次不得改变", "active.json"),
        ("V2.3 包内容哈希", manifest["content_sha256"], "不可变包校验", "2.3.0 manifest.json"),
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    for row_number in (5, 7):
        summary.cell(row_number, 1).fill = WARNING_FILL
        summary.cell(row_number, 2).fill = WARNING_FILL
    summary["B16"].comment = Comment(
        "候选包未激活；active.json 仍指向 2.1.0。",
        "GIS Data Agent",
    )

    _write_sheet(
        workbook,
        "清单逐项覆盖",
        list(coverage["coverage"]),
        [
            ("sequence", "序号"),
            ("sheet", "原工作表"),
            ("excel_row", "原行号"),
            ("data_category", "数据类别"),
            ("entity_code", "实体编码"),
            ("entity_label", "实体名称"),
            ("source_layer", "来源图层"),
            ("source_layer_code", "来源图层代码"),
            ("source_data", "来源数据"),
            ("source_attribute", "清单属性"),
            ("source_field_code", "清单字段代码"),
            ("target_domain_class", "目标领域类"),
            ("target_data_property", "目标领域属性"),
            ("target_object_relation", "目标对象关系"),
            ("v2_2_coverage_status", "最低覆盖状态"),
            ("modeling_note", "建模备注"),
        ],
        "AttachmentCoverage",
    )
    _write_sheet(
        workbook,
        "图层证据匹配",
        list(audit["layers"]),
        [
            ("layer_number", "序号"),
            ("source_layer", "清单来源图层"),
            ("source_layer_code", "清单来源代码"),
            ("source_data", "来源数据"),
            ("entity_codes", "实体编码"),
            ("entity_labels", "实体名称"),
            ("attachment_field_count", "清单字段数"),
            ("lookup_codes", "检索代码"),
            ("target_class", "清单目标类"),
            ("status", "证据状态"),
            ("schema_candidate_count", "命中结构数"),
            ("expanded_field_count", "展开字段数"),
            ("bound_target_classes", "绑定领域类"),
            ("schema_candidates", "EA/标准结构证据"),
        ],
        "LayerEvidence",
    )
    _write_sheet(
        workbook,
        "完整字段语义处置",
        list(audit["expanded_fields"]),
        [
            ("layer_number", "图层序号"),
            ("attachment_layer", "清单来源图层"),
            ("attachment_layer_code", "清单来源代码"),
            ("target_class", "清单目标类"),
            ("schema_concept_id", "结构制品 ID"),
            ("schema_source_id", "结构来源"),
            ("schema_code", "结构代码"),
            ("schema_label", "结构名称"),
            ("source_property_id", "来源字段 ID"),
            ("field_code", "字段代码"),
            ("field_label", "字段名称"),
            ("datatype", "数据类型"),
            ("required", "必填"),
            ("value_domain", "值域"),
            ("target_property", "目标领域属性"),
            ("target_relation", "目标对象关系"),
            ("semantic_status", "语义处置"),
            ("target_class_ids", "语义目标类"),
            ("semantic_mapping_basis", "映射依据"),
            ("semantic_exclusion_reason", "排除原因"),
        ],
        "ExpandedFields",
    )
    _write_sheet(
        workbook,
        "领域类",
        class_rows,
        [
            ("code", "类代码"),
            ("label", "类名称"),
            ("kind", "建模角色"),
            ("parent_code", "父类代码"),
            ("parent_label", "父类名称"),
            ("domain_id", "主题域"),
            ("definition", "定义"),
            ("aliases", "别名"),
            ("evidence_status", "证据状态"),
            ("evidence_count", "来源证据数"),
        ],
        "DomainClasses",
    )
    _write_sheet(
        workbook,
        "领域属性",
        property_rows,
        [
            ("code", "属性代码"),
            ("label", "属性名称"),
            ("owner_code", "所属类代码"),
            ("owner_label", "所属类名称"),
            ("datatype", "数据类型"),
            ("min_count", "最小基数"),
            ("max_count", "最大基数"),
            ("source_codes", "来源字段代码"),
            ("source_labels", "来源字段名称"),
            ("definition", "定义"),
        ],
        "DomainProperties",
    )
    _write_sheet(
        workbook,
        "对象关系",
        relation_rows,
        [
            ("name", "关系代码"),
            ("label", "关系名称"),
            ("domain_code", "定义域代码"),
            ("domain_label", "定义域名称"),
            ("range_code", "值域代码"),
            ("range_label", "值域名称"),
            ("inverse", "逆关系"),
            ("functional", "函数属性"),
            ("restriction", "存在性约束"),
        ],
        "ObjectRelations",
    )
    gaps = [row for row in audit["layers"] if row["status"] != "matched"]
    _write_sheet(
        workbook,
        "证据缺口",
        gaps,
        [
            ("source_layer", "清单来源图层"),
            ("source_layer_code", "清单来源代码"),
            ("source_data", "来源数据"),
            ("entity_codes", "实体编码"),
            ("entity_labels", "实体名称"),
            ("attachment_field_count", "清单字段数"),
            ("target_class", "最低目标类"),
            ("status", "缺口状态"),
        ],
        "EvidenceGaps",
    )
    boundary_rows = [
        {
            "term": "自然资源领域本体",
            "definition": "现实领域概念、关系、公理与约束的形式模型",
            "excluded": "数据表目录、字段全集、智能问数功能",
        },
        {
            "term": "领域类",
            "definition": "可用于实例归类和推理的稳定自然资源概念",
            "excluded": "表、图层、字段、EA package、菜单功能",
        },
        {
            "term": "领域属性",
            "definition": "含义独立于存储字段命名的稳定实例特征",
            "excluded": "数据库列或标准表字段本身",
        },
        {
            "term": "对象关系",
            "definition": "连接两个领域对象的权利、空间、组成、观测、登记或证据语义",
            "excluded": "外键字符串",
        },
        {
            "term": "来源 Schema 制品",
            "definition": "EA 表、标准属性结构表或数据集结构等来源证据",
            "excluded": "领域 OWL 类",
        },
        {
            "term": "来源字段",
            "definition": "隶属于来源 Schema 的字段，可映射但本身不是领域属性",
            "excluded": "自动提升为 OWL DatatypeProperty",
        },
        {
            "term": "清单覆盖基线",
            "definition": "附件定义的最低逐项验收范围",
            "excluded": "完整本体或唯一规范来源",
        },
        {
            "term": "证据缺口",
            "definition": "冻结 EA/标准中没有可核验完整结构的状态",
            "excluded": "推测或伪造字段",
        },
    ]
    _write_sheet(
        workbook,
        "建模边界",
        boundary_rows,
        [("term", "术语"), ("definition", "定义"), ("excluded", "明确排除")],
        "ModelBoundary",
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--coverage", type=Path, default=DEFAULT_COVERAGE)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--package", type=Path, default=DEFAULT_PACKAGE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    build(args.coverage, args.audit, args.package, args.output)
    print(args.output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
